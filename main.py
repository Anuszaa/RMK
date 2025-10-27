# -*- coding: utf-8 -*-
"""
RMK insGT - Aplikacja do zarządzania rozliczeniami międzyokresowymi kosztów
Obsługuje polskie znaki i formatowanie liczb zgodnie z polską konwencją

LOGO FIRMY:
Aby dodać logo firmy, umieść plik obrazu (.png, .jpg, .jpeg) w jednej z lokalizacji:
1. Folder aplikacji/logo/ (zalecane)
2. Główny folder aplikacji  
3. Twój folder domyślny (~)
4. Pulpit (Desktop)
5. Dokumenty (Documents)

Obsługiwane nazwy plików logo:
- RMK_insGT_logo.png, RMK insG.png
- logo.png, logo.jpg, logo.jpeg
- rmk.png, rmk.jpg, rmk.jpeg
- company_logo.png, firma_logo.png
- lub dowolny plik z "logo" w nazwie

Logo zostanie automatycznie przeskalowane i wyświetlone w lewym górnym rogu aplikacji.
"""

import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
from datetime import datetime, date, timedelta
import traceback
from dataclasses import dataclass
from typing import List, Optional, Dict
import sys, os, json
import re

# Poprawka dla PyInstaller - określenie base path
def resource_path(relative_path):
    """Pobierz ścieżkę do zasobów, działając zarówno w dev jak i w PyInstaller"""
    try:
        # PyInstaller tworzy folder tymczasowy i przechowuje ścieżkę w _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def create_btn(owner, parent, **kwargs):
    """Create a button using owner's _btn if available (owner may be RMKApp or a dialog).
    Falls back to ttk.Button and maps bootstyle -> style if needed.
    """
    # If owner itself provides _btn (RMKApp methods), use it
    try:
        if hasattr(owner, '_btn') and callable(getattr(owner, '_btn')):
            return owner._btn(parent, **kwargs)
    except Exception:
        pass
    # If owner has a master (dialogs), try master's _btn
    try:
        m = getattr(owner, 'master', None)
        if m and hasattr(m, '_btn') and callable(getattr(m, '_btn')):
            return m._btn(parent, **kwargs)
    except Exception:
        pass
    # fallback: map bootstyle to style and create plain ttk.Button
    bootstyle = kwargs.pop('bootstyle', None)
    style = kwargs.pop('style', None)
    if not style and bootstyle:
        # Mapowanie bootstyle na nasze zdefiniowane style
        style_map = {
            'success': 'Success.TButton',
            'warning': 'Warning.TButton', 
            'danger': 'Danger.TButton',
            'primary': 'Info.TButton',
            'info': 'Info.TButton'
        }
        style = style_map.get(bootstyle, 'TButton')
    elif not style:
        style = 'TButton'
    return ttk.Button(parent, style=style, **kwargs)

# optional dependency for Excel import
try:
    import openpyxl
except Exception:
    openpyxl = None
# optional PDF export dependency
try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.utils import ImageReader
    from reportlab.lib.styles import getSampleStyleSheet
    REPORTLAB_AVAILABLE = True
except Exception:
    REPORTLAB_AVAILABLE = False

# optional ttkbootstrap for nicer native styles
try:
    from ttkbootstrap import Style as TBStyle
    TTKBOOTSTRAP_AVAILABLE = True
except Exception:
    TBStyle = None
    TTKBOOTSTRAP_AVAILABLE = False

APP_NAME = "RMK insGT"
APP_VERSION = "v0.22.15"
COMPANY_NAME = "IntegritasAD"

# Kolorystyka w stylu SAP
BRAND_COLOR_BG = "#003366"          # Ciemnoniebieski SAP
BRAND_COLOR_ACCENT = "#0070BA"      # Niebieski SAP
BRAND_COLOR_LIGHT = "#AFAFAF"       # Ciemniejsze szare tło główne SAP
BRAND_COLOR_FRAME = "#838181"       # Lekko szare panele robocze SAP zamiast białych
BRAND_COLOR_BORDER = "#CCCCCC"      # Szare obramowania SAP
BRAND_COLOR_HEADER = "#E8E8E8"      # Jasny szary dla nagłówków SAP
BRAND_COLOR_YELLOW = "#FFD700"      # Żółty akcent SAP
BRAND_COLOR_TEXT = "#333333"        # Ciemnoszary tekst SAP

def configure_ttk_styles(root):
    """Konfiguruje style TTK dla zachowania kolorów w exe"""
    try:
        # Debug kolorów SAP
        print("DEBUG KOLORÓW SAP:")
        print(f"   BRAND_COLOR_BG = {BRAND_COLOR_BG}")
        print(f"   BRAND_COLOR_ACCENT = {BRAND_COLOR_ACCENT}")
        print(f"   BRAND_COLOR_LIGHT = {BRAND_COLOR_LIGHT}")
        print(f"   BRAND_COLOR_FRAME = {BRAND_COLOR_FRAME}")
        print(f"   BRAND_COLOR_HEADER = {BRAND_COLOR_HEADER}")
        
        style = ttk.Style(root)
        
        # Wymuś nasz motyw i style - zastąp ttkbootstrap
        try:
            # Sprawdź dostępne motywy i wybierz najlepszy
            available_themes = style.theme_names()
            print(f"Dostępne motywy TTK: {available_themes}")
            
            if 'clam' in available_themes:
                style.theme_use('clam')  # Bardziej nowoczesny wygląd
                print("Używam motywu 'clam'")
            elif 'alt' in available_themes:
                style.theme_use('alt')   # Alternatywny motyw
                print("Używam motywu 'alt'")
            else:
                style.theme_use('default')
                print("Używam motywu 'default'")
        except Exception as e:
            print(f"BŁĄD motywu: {e}")
            style.theme_use('default')
        
        # Konfiguruj style dla lepszego wyglądu
        # Style SAP - etykiety
        style.configure('TLabel', 
                       background=BRAND_COLOR_LIGHT,
                       foreground=BRAND_COLOR_TEXT,
                       font=('Segoe UI', 9))
        
        # Style SAP - ramki
        style.configure('TFrame', 
                       background=BRAND_COLOR_FRAME,  # Białe panele SAP
                       borderwidth=1,
                       relief='solid',
                       bordercolor=BRAND_COLOR_BORDER)  # Szare obramowania SAP
        
        # Specjalny styl dla zakładek - jasnoszare tło SAP
        style.configure('Light.TFrame', 
                       background=BRAND_COLOR_LIGHT,  # Główne tło SAP
                       borderwidth=0,
                       relief='flat')
        
        # Styl dla sekcji/paneli - białe panele robocze SAP
        style.configure('Panel.TFrame', 
                       background=BRAND_COLOR_FRAME,
                       borderwidth=1,
                       relief='solid',
                       bordercolor=BRAND_COLOR_BORDER)
        
        # Style SAP - pola wprowadzania
        style.configure('TEntry', 
                       fieldbackground='white',
                       borderwidth=1,
                       relief='solid',
                       bordercolor=BRAND_COLOR_BORDER,  # Szare obramowania SAP
                       font=('Segoe UI', 9))
        
        style.configure('TCombobox', 
                       fieldbackground='white',
                       borderwidth=1,
                       relief='solid',
                       bordercolor=BRAND_COLOR_BORDER,  # Szare obramowania SAP
                       font=('Segoe UI', 9))
        
        # Style SAP - przyciski główne
        style.configure('TButton',
                       background=BRAND_COLOR_ACCENT,  # Niebieski SAP
                       foreground='white',
                       borderwidth=1,
                       focuscolor='none',
                       font=('Segoe UI', 9, 'bold'),
                       relief='raised')
        
        style.map('TButton',
                 background=[('active', BRAND_COLOR_BG),  # Ciemnoniebieski SAP przy hover
                            ('pressed', BRAND_COLOR_BG)],
                 relief=[('pressed', 'sunken')])
        
        # Style dla przycisków o różnych kolorach - bardziej wyraziste
        style.configure('Success.TButton',
                       background='#28a745',
                       foreground='white',
                       borderwidth=2,
                       relief='raised',
                       font=('Segoe UI', 9, 'bold'))
        
        style.configure('Warning.TButton',
                       background='#ffc107',
                       foreground='black',
                       borderwidth=2,
                       relief='raised',
                       font=('Segoe UI', 9, 'bold'))
        
        style.configure('Danger.TButton',
                       background='#dc3545',
                       foreground='white',
                       borderwidth=2,
                       relief='raised',
                       font=('Segoe UI', 9, 'bold'))
        
        style.configure('Info.TButton',
                       background='#17a2b8',
                       foreground='white',
                       borderwidth=2,
                       relief='raised',
                       font=('Segoe UI', 9, 'bold'))
        
        # Style SAP - tabele z alternującymi wierszami
        style.configure('Treeview',
                       background='#FFFFFF',            # Białe tło tabel SAP
                       foreground=BRAND_COLOR_TEXT,     # Ciemnoszary tekst SAP
                       rowheight=25,
                       fieldbackground='#FFFFFF',       # Białe tło pól
                       font=('Segoe UI', 9),
                       borderwidth=1,
                       relief='solid',
                       bordercolor=BRAND_COLOR_BORDER,  # Szare ramki SAP
                       lightcolor=BRAND_COLOR_BORDER,   # Separatory kolumn SAP
                       darkcolor='#B0B0B0')             # Separatory wierszy SAP
        
        # Style specjalne - tabele Słowników
        style.configure('Dictionary.Treeview',
                       background='#FFFFFF',
                       foreground=BRAND_COLOR_TEXT,
                       rowheight=25,
                       fieldbackground='#FFFFFF',
                       font=('Segoe UI', 9),
                       borderwidth=1,
                       relief='solid',
                       bordercolor=BRAND_COLOR_BORDER,
                       lightcolor=BRAND_COLOR_BORDER,
                       darkcolor='#B0B0B0')
        
        # Style specjalne - tabele Admin
        style.configure('Admin.Treeview',
                       background='#FFFFFF',
                       foreground=BRAND_COLOR_TEXT,
                       rowheight=25,
                       fieldbackground='#FFFFFF',
                       font=('Segoe UI', 9),
                       borderwidth=1,
                       relief='solid',
                       bordercolor=BRAND_COLOR_BORDER,
                       lightcolor=BRAND_COLOR_BORDER,
                       darkcolor='#B0B0B0')
        
        # Style SAP - nagłówki tabel (niebieski SAP) - domyślne
        style.configure('Treeview.Heading',
                       background=BRAND_COLOR_ACCENT,   # Niebieski SAP
                       foreground='white',
                       font=('Segoe UI', 9, 'bold'),
                       borderwidth=1,
                       relief='raised',
                       bordercolor=BRAND_COLOR_BORDER,
                       lightcolor='#ffffff',
                       darkcolor='#666666')
        
        # Style specjalne - nagłówki Słowników (bold)
        style.configure('Dictionary.Treeview.Heading',
                       background=BRAND_COLOR_ACCENT,   # Niebieski SAP
                       foreground='white',
                       font=('Segoe UI', 10, 'bold'),  # Większy i bold
                       borderwidth=1,
                       relief='raised',
                       bordercolor=BRAND_COLOR_BORDER,
                       lightcolor='#ffffff',
                       darkcolor='#666666')
        
        # Style specjalne - nagłówki Admin (normal)
        style.configure('Admin.Treeview.Heading',
                       background=BRAND_COLOR_ACCENT,   # Niebieski SAP
                       foreground='white',
                       font=('Segoe UI', 9),            # Normal weight
                       borderwidth=1,
                       relief='raised',
                       bordercolor=BRAND_COLOR_BORDER,
                       lightcolor='#ffffff',
                       darkcolor='#666666')
        
        # Mapowanie SAP - ciemnoniebieski przy hover
        style.map('Treeview.Heading',
                 background=[('active', BRAND_COLOR_BG)],  # Ciemnoniebieski SAP
                 relief=[('active', 'raised')])
        
        style.map('Treeview',
                 background=[('selected', BRAND_COLOR_ACCENT)],
                 foreground=[('selected', 'white')])
        
        # Element dla obramowań wierszy i kolumn
        style.configure('Treeview.Item',
                       borderwidth=1,
                       relief='solid',
                       bordercolor='#cccccc')
        
        style.configure('Treeview.Cell',
                       borderwidth=1,
                       relief='solid',
                       bordercolor='#cccccc')
        
        # Style SAP - selekcja w tabelach (niebieski SAP)
        style.map('Treeview',
                 background=[('selected', BRAND_COLOR_ACCENT)],  # Niebieski SAP
                 foreground=[('selected', 'white')])
        
        # Style SAP - zakładki (Notebook)
        style.configure('TNotebook',
                       background=BRAND_COLOR_LIGHT,    # Jasnoszare tło SAP
                       borderwidth=0)
        
        style.configure('TNotebook.Tab',
                       background="#E0E0E0",            # Szare nieaktywne zakładki SAP
                       foreground=BRAND_COLOR_TEXT,
                       padding=[8, 4],
                       font=('Segoe UI', 9))
        
        style.map('TNotebook.Tab',
                 background=[('selected', BRAND_COLOR_FRAME),  # Biała aktywna zakładka SAP
                            ('active', BRAND_COLOR_ACCENT)],    # Niebieski SAP przy hover
                 foreground=[('selected', BRAND_COLOR_TEXT),   # Ciemnoszary tekst SAP
                            ('active', 'white')])
        
        # Style SAP - LabelFrame (sekcje)
        style.configure('TLabelframe',
                       background=BRAND_COLOR_FRAME,     # Białe tło SAP
                       borderwidth=1,
                       relief='solid',
                       bordercolor=BRAND_COLOR_BORDER)   # Szare obramowania SAP
        
        style.configure('TLabelframe.Label',
                       background=BRAND_COLOR_FRAME,     # Białe tło SAP
                       foreground=BRAND_COLOR_TEXT,      # Ciemnoszary tekst SAP
                       font=('Segoe UI', 9, 'bold'))
        
        print("OK Style TTK skonfigurowane pomyślnie")
        
        # Zaplanuj konfigurację tagów Treeview po utworzeniu wszystkich widżetów
        root.after(100, lambda: configure_treeview_tags(root))
        root.after(200, lambda: configure_treeview_borders(root))  # Dodaj obramowania
        
    except Exception as e:
        print(f"UWAGA  Błąd konfiguracji stylów TTK: {e}")

def configure_single_treeview_borders(tree):
    """Konfiguruj obramowania dla pojedynczej tabeli"""
    try:
        if hasattr(tree, '_rmk_bordered'):
            return  # Już skonfigurowane
            
        print(f"🔲 Konfiguruję obramowania dla tabeli: {tree}")
        
        # Dla ttk.Treeview używaj tylko obsługiwanych opcji
        try:
            tree.configure(
                selectmode='extended'
            )
            
            # Wymuś separatory przez style SAP
            style = ttk.Style()
            style.configure('Bordered.Treeview',
                          background='white',                  # Białe tło SAP
                          foreground=BRAND_COLOR_TEXT,         # Ciemnoszary tekst SAP
                          borderwidth=1,
                          relief='solid',
                          bordercolor=BRAND_COLOR_BORDER,      # Szare obramowania SAP
                          lightcolor=BRAND_COLOR_BORDER,       # Separatory kolumn SAP
                          darkcolor='#B0B0B0')                 # Separatory wierszy SAP
            
            style.configure('Bordered.Treeview.Heading',
                          background=BRAND_COLOR_ACCENT,       # Niebieski SAP
                          foreground='white',
                          font=('Segoe UI', 9, 'bold'),
                          borderwidth=1,
                          relief='raised',
                          bordercolor=BRAND_COLOR_BORDER)
            
            # Zastosuj styl do tabeli
            tree.configure(style='Bordered.Treeview')
            
        except Exception as e:
            print(f"UWAGA Błąd konfiguracji podstawowej: {e}")
        
        # Definiuj tagi z konkretnymi kolorami
        tree.tag_configure('evenrow', background='#f8f9fa', foreground='black')
        tree.tag_configure('oddrow', background='white', foreground='black')
        
        # Hook dla kolorowania wierszy
        def colorize_rows(*args):
            try:
                items = tree.get_children()
                for i, item in enumerate(items):
                    current_tags = list(tree.item(item, 'tags'))
                    
                    # Zachowaj ważne tagi statusu (gen, ungen)
                    important_tags = [tag for tag in current_tags if tag in ['gen', 'ungen']]
                    
                    # Jeśli ma ważny tag statusu, zachowaj go bez zmian
                    if important_tags:
                        continue
                    
                    # Dla pozostałych stosuj alternujące kolory
                    tag = 'evenrow' if i % 2 == 0 else 'oddrow'
                    tree.item(item, tags=[tag])
            except Exception as e:
                print(f"Błąd kolorowania: {e}")
        
        # Binduj zdarzenia
        tree.bind('<Map>', lambda e: colorize_rows())
        tree.bind('<<TreeviewSelect>>', lambda e: tree.after_idle(colorize_rows))
        
        tree._rmk_bordered = True
        
        # Zastosuj kolory natychmiast
        colorize_rows()
        
        print(f"OK Obramowania skonfigurowane dla tabeli (tylko alternujące kolory)")
        
    except Exception as e:
        print(f"UWAGA Błąd obramowań tabeli: {e}")

def configure_treeview_borders(root):
    """Dodaj faktyczne linie obramowania do tabel"""
    try:
        def apply_real_borders(widget):
            if isinstance(widget, ttk.Treeview):
                print(f"🔲 Dodaję RZECZYWISTE obramowania dla tabeli")
                
                try:
                    # 1. Konfiguruj samą tabelę
                    widget.configure(
                        selectmode='extended',
                        background='white',
                        foreground='black', 
                        fieldbackground='white'
                    )
                    
                    # 2. Alternujące kolory wierszy
                    widget.tag_configure('evenrow', background='#f0f8ff', foreground='black')
                    widget.tag_configure('oddrow', background='white', foreground='black')
                    
                    def apply_row_colors():
                        children = widget.get_children()
                        for i, child in enumerate(children):
                            tag = 'evenrow' if i % 2 == 0 else 'oddrow'
                            widget.item(child, tags=[tag])
                    
                    # 3. Hook tylko jeśli jeszcze nie ma
                    if not hasattr(widget, '_rmk_bordered'):
                        original_insert = widget.insert
                        def new_insert(*args, **kwargs):
                            result = original_insert(*args, **kwargs)
                            widget.after_idle(apply_row_colors)
                            return result
                        widget.insert = new_insert
                        
                        original_delete = widget.delete  
                        def new_delete(*args, **kwargs):
                            result = original_delete(*args, **kwargs)
                            widget.after_idle(apply_row_colors)
                            return result
                        widget.delete = new_delete
                        
                        widget._rmk_bordered = True
                    
                    # 4. Dodaj Canvas z liniami
                    def add_border_canvas():
                        try:
                            parent = widget.master
                            
                            # Znajdź pozycję tabeli
                            x = widget.winfo_x()
                            y = widget.winfo_y()
                            width = widget.winfo_width()
                            height = widget.winfo_height()
                            
                            # Stwórz Canvas dla linii
                            canvas = tk.Canvas(parent, 
                                             width=width, height=height,
                                             highlightthickness=0, bd=0,
                                             bg='', takefocus=False)
                            canvas.place(x=x, y=y)
                            
                            def draw_grid():
                                try:
                                    canvas.delete("border_line")
                                    
                                    # Pobierz aktualne wymiary
                                    w = widget.winfo_width()
                                    h = widget.winfo_height()
                                    
                                    if w > 10 and h > 10:
                                        # Linie poziome (wiersze)
                                        row_height = 25  # Wysokość wiersza
                                        y_pos = row_height  # Zaczynaj po nagłówku
                                        while y_pos < h:
                                            canvas.create_line(0, y_pos, w, y_pos, 
                                                             fill='#cccccc', width=1, 
                                                             tags="border_line")
                                            y_pos += row_height
                                        
                                        # Linie pionowe (kolumny)
                                        if hasattr(widget, 'cget') and widget.cget('columns'):
                                            cols = widget.cget('columns')
                                            x_pos = 0
                                            for col in cols:
                                                try:
                                                    col_width = widget.column(col, 'width')
                                                    x_pos += col_width
                                                    if x_pos < w:
                                                        canvas.create_line(x_pos, 0, x_pos, h,
                                                                         fill='#cccccc', width=1,
                                                                         tags="border_line")
                                                except:
                                                    pass
                                        
                                        # Obramowanie zewnętrzne
                                        canvas.create_rectangle(0, 0, w-1, h-1,
                                                              outline='#999999', width=1,
                                                              tags="border_line")
                                        
                                        print(f"OK Narysowano linie: {w}x{h}")
                                        
                                except Exception as e:
                                    print(f"UWAGA Błąd rysowania linii: {e}")
                            
                            # Aktualizuj po zmianach
                            widget.bind('<Configure>', lambda e: root.after_idle(draw_grid))
                            widget.bind('<Map>', lambda e: root.after_idle(draw_grid))
                            
                            # Pierwsze rysowanie
                            root.after(500, draw_grid)
                            
                        except Exception as e:
                            print(f"UWAGA Błąd Canvas overlay: {e}")
                    
                    # Uruchom dodawanie Canvas z opóźnieniem
                    root.after(300, add_border_canvas)
                    
                    # 5. Zastosuj kolory natychmiast
                    apply_row_colors()
                    
                    print(f"OK Rzeczywiste obramowania dodane")
                    
                except Exception as e:
                    print(f"UWAGA Błąd obramowań: {e}")
            
            # Rekurencyjnie
            for child in widget.winfo_children():
                apply_real_borders(child)
        
        apply_real_borders(root)
        print("OK Rzeczywiste obramowania tabel skonfigurowane")
        
    except Exception as e:
        print(f"UWAGA Błąd konfiguracji obramowań: {e}")

def configure_treeview_tags(root):
    """Konfiguruje tagi kolorystyczne dla wszystkich Treeview w aplikacji"""
    try:
        def configure_widget_tags(widget):
            if isinstance(widget, ttk.Treeview):
                # Alternujące kolory wierszy - bardziej widoczne
                widget.tag_configure('odd', 
                                   background='#F8F8F8',   # Jasnoszary
                                   foreground=BRAND_COLOR_TEXT)
                widget.tag_configure('even', 
                                   background='#FFFFFF',   # Biały
                                   foreground=BRAND_COLOR_TEXT)
                
                # Kolory statusu - bardziej wyraziste
                widget.tag_configure('gen', 
                                   background='#d4edda',   # Jasnozielony
                                   foreground='#155724',   # Ciemnozielony
                                   font=('Segoe UI', 9))
                widget.tag_configure('ungen', 
                                   background='#f8d7da',   # Jasnoróżowy
                                   foreground='#721c24',   # Ciemnoczerwony  
                                   font=('Segoe UI', 9))
                
                # Dodaj automatyczne kolorowanie wierszy
                def apply_alternating_colors():
                    """Stosuje alternujące kolory do wszystkich wierszy, zachowując kolory statusu"""
                    children = widget.get_children()
                    for i, child in enumerate(children):
                        current_tags = list(widget.item(child, 'tags'))
                        
                        # Zachowaj ważne tagi statusu
                        important_tags = [tag for tag in current_tags if tag in ['gen', 'ungen']]
                        
                        # Jeśli ma ważny tag statusu, nie zmieniaj kolorów
                        if important_tags:
                            continue
                            
                        # Dla pozostałych wierszy stosuj alternujące kolory
                        if i % 2 == 0:
                            # Usuń stare tagi kolorów alternujących
                            tags = [tag for tag in current_tags if tag not in ['odd', 'even']]
                            tags.append('even')
                            widget.item(child, tags=tags)
                        else:
                            # Usuń stare tagi kolorów alternujących
                            tags = [tag for tag in current_tags if tag not in ['odd', 'even']]
                            tags.append('odd')
                            widget.item(child, tags=tags)
                
                # Zastosuj kolory przy zmianie zawartości
                widget.bind('<<TreeviewOpen>>', lambda e: widget.after(10, apply_alternating_colors))
                widget.bind('<<TreeviewClose>>', lambda e: widget.after(10, apply_alternating_colors))
                
                # Pierwsze zastosowanie
                widget.after(100, apply_alternating_colors)
            
            # Rekurencyjnie przeglądaj dzieci
            for child in widget.winfo_children():
                configure_widget_tags(child)
        
        configure_widget_tags(root)
        print("OK Tagi Treeview skonfigurowane pomyślnie")
        
    except Exception as e:
        print(f"UWAGA  Błąd konfiguracji tagów Treeview: {e}")

@dataclass
class RMKItem:
    id: int
    opis: str
    data_start: date
    liczba_mies: int
    kwota: float
    firma: str
    kategoria: str
    konto_kosztowe: str
    konto_rmk: str
    numer_faktury: str
    kontrahent: str
    uwagi: str = ""
    data_koniec: Optional[date] = None
    harmonogram_generated: bool = False
    harmonogram: Optional[List[Dict]] = None
    def to_dict(self):
        return {
            'id': self.id,
            'opis': self.opis,
            'data_start': self.data_start.isoformat(),
            'liczba_mies': self.liczba_mies,
            'kwota': self.kwota,
            'firma': self.firma,
            'kategoria': self.kategoria,
            'konto_kosztowe': self.konto_kosztowe,
            'konto_rmk': self.konto_rmk,
            'numer_faktury': self.numer_faktury,
            'kontrahent': self.kontrahent,
            'uwagi': self.uwagi,
            'data_koniec': self.data_koniec.isoformat() if self.data_koniec else None,
            'harmonogram_generated': bool(self.harmonogram_generated),
            'harmonogram': self.harmonogram or []
        }
    @staticmethod
    def from_dict(d):
        return RMKItem(
            id=int(d['id']),
            opis=d.get('opis',''),
            data_start=datetime.fromisoformat(d['data_start']).date(),
            liczba_mies=int(d.get('liczba_mies',0)),
            kwota=float(d.get('kwota',0.0)),
            firma=d.get('firma',''),
            kategoria=d.get('kategoria',''),
            konto_kosztowe=d.get('konto_kosztowe',''),
            konto_rmk=d.get('konto_rmk',''),
            numer_faktury=d.get('numer_faktury',''),
            kontrahent=d.get('kontrahent',''),
            uwagi=d.get('uwagi',''),
            data_koniec=datetime.fromisoformat(d['data_koniec']).date() if d.get('data_koniec') else None,
            harmonogram_generated=bool(d.get('harmonogram_generated', False)),
            harmonogram=d.get('harmonogram', [])
        )

def thousand_sep(value: float) -> str:
    """Format liczby: grupowanie tysięcy spacją, przecinek jako separator dziesiętny.

    Przykład: 1234.56 -> '1 234,56'
    """
    s = f"{value:,.2f}"  # '1,234.56' for en locale
    s = s.replace(',', ' ')  # thousands -> space
    s = s.replace('.', ',')  # decimal point -> comma
    return s

def month_add(dt: date, months: int) -> date:
    m = dt.month - 1 + months
    y = dt.year + m // 12
    m = m % 12 + 1
    d = min(dt.day, [31, 29 if y % 4 == 0 and (y % 100 != 0 or y % 400 == 0) else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31][m - 1])
    return date(y, m, d)

def calculate_monthly_amounts_improved(total_amount: float, start_date: date, num_months: int):
    """
    Ulepszone rozliczanie miesięczne:
    - Pierwszy miesiąc: proporcjonalnie do liczby dni
    - Ostatni miesiąc: proporcjonalnie do liczby dni
    - Środkowe miesiące: stała kwota
    """
    if num_months <= 0:
        return []
    
    if num_months == 1:
        # Tylko jeden miesiąc - zwróć całą kwotę
        return [total_amount]
    
    amounts = []
    remaining_amount = total_amount
    
    # Pierwszy miesiąc - proporcjonalnie do dni
    first_month_date = start_date
    first_month_days_total = (date(first_month_date.year, first_month_date.month + 1, 1) - date(first_month_date.year, first_month_date.month, 1)).days if first_month_date.month < 12 else (date(first_month_date.year + 1, 1, 1) - date(first_month_date.year, first_month_date.month, 1)).days
    first_month_days_used = first_month_days_total - first_month_date.day + 1
    
    # Ostatni miesiąc - znajdź datę końcową
    end_date = month_add(start_date, num_months - 1)
    last_month_days_total = (date(end_date.year, end_date.month + 1, 1) - date(end_date.year, end_date.month, 1)).days if end_date.month < 12 else (date(end_date.year + 1, 1, 1) - date(end_date.year, end_date.month, 1)).days
    last_month_days_used = end_date.day
    
    if num_months == 2:
        # Tylko dwa miesiące - podziel proporcjonalnie
        total_days = first_month_days_used + last_month_days_used
        first_amount = round(total_amount * first_month_days_used / total_days, 2)
        last_amount = round(total_amount - first_amount, 2)
        return [first_amount, last_amount]
    
    # Więcej niż 2 miesiące
    # Oblicz średnią miesięczną kwotę dla pełnych miesięcy środkowych
    middle_months = num_months - 2
    
    # Estymuj udział pierwszego i ostatniego miesiąca
    first_ratio = first_month_days_used / first_month_days_total
    last_ratio = last_month_days_used / last_month_days_total
    
    # Oblicz kwotę dla pełnego miesiąca
    full_month_amount = total_amount / (first_ratio + middle_months + last_ratio)
    
    # Pierwszy miesiąc
    first_amount = round(full_month_amount * first_ratio, 2)
    amounts.append(first_amount)
    remaining_amount -= first_amount
    
    # Środkowe miesiące - stała kwota
    middle_amount = round(full_month_amount, 2)
    for i in range(middle_months):
        if i == middle_months - 1:  # Ostatni ze środkowych miesięcy
            # Zostaw miejsce na ostatni miesiąc
            last_amount_estimate = round(full_month_amount * last_ratio, 2)
            middle_this = round(remaining_amount - last_amount_estimate, 2)
        else:
            middle_this = middle_amount
        amounts.append(middle_this)
        remaining_amount -= middle_this
    
    # Ostatni miesiąc - reszta kwoty
    amounts.append(round(remaining_amount, 2))
    
    return amounts

class Splash(tk.Toplevel):
    def __init__(self, master):
        super().__init__(master)
        print(f"🌟 Debug splash - tworzę splash screen")
        print(f"🌟 Debug splash - kolory: BG={BRAND_COLOR_BG}, YELLOW={BRAND_COLOR_YELLOW}")
        
        # Wymuś kolory dla exe
        bg_color = BRAND_COLOR_BG or "#0F2D52"  # Fallback
        yellow_color = BRAND_COLOR_YELLOW or "#FFEB80"  # Fallback
        
        self.overrideredirect(True)
        self.configure(bg=bg_color)
        w, h = 420, 240
        x = master.winfo_screenwidth() // 2 - w // 2
        y = master.winfo_screenheight() // 2 - h // 2
        self.geometry(f"{w}x{h}+{x}+{y}")
        
        print(f"🌟 Debug splash - geometria: {w}x{h}+{x}+{y}")
        print(f"🌟 Debug splash - używam kolorów: bg={bg_color}, fg={yellow_color}")
        
        # Wymuś kolory bezpośrednio - żółte litery bez dodatkowego tła
        company_label = tk.Label(self, text=COMPANY_NAME, fg=yellow_color, bg=bg_color, 
                                font=("Segoe UI", 22, "bold"))
        company_label.pack(pady=(40, 10))
        
        # Dodatkowe formatowanie żeby usunąć tło wokół liter
        company_label.configure(relief='flat', bd=0, highlightthickness=0)
        
        tk.Label(self, text=APP_NAME, fg="white", bg=bg_color, font=("Segoe UI", 16)).pack()
        tk.Label(self, text="Wczytywanie...", fg="#D0E3FF", bg=bg_color, font=("Segoe UI", 10)).pack(pady=20)
        
        print(f"🌟 Debug splash - etykiety utworzone")
        self.after(1200, self.destroy)

class LoginDialog(tk.Toplevel):
    def __init__(self, master, attempts_left: Optional[int] = None):
        super().__init__(master)
        self.title("Logowanie")
        self.configure(bg=BRAND_COLOR_LIGHT)
        self.resizable(False, False)
        w, h = 420, 240
        x = self.winfo_screenwidth() // 2 - w // 2
        y = self.winfo_screenheight() // 2 - h // 2
        self.geometry(f"{w}x{h}+{x}+{y}")
        try:
            if getattr(master, 'winfo_ismapped', lambda: False)() or getattr(master, 'state', lambda: '')() == 'normal':
                try:
                    self.transient(master)
                except Exception:
                    pass
            try:
                self.grab_set()
            except Exception:
                pass
        except Exception:
            pass
        self.protocol("WM_DELETE_WINDOW", self.cancel)

        frame = ttk.Frame(self, padding=16)
        frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frame, text="Użytkownik:").grid(row=0, column=0, sticky="e", pady=6, padx=6)
        ttk.Label(frame, text="Hasło:").grid(row=1, column=0, sticky="e", pady=6, padx=6)
        ttk.Label(frame, text="Firma:").grid(row=2, column=0, sticky="e", pady=6, padx=6)

        self.user = ttk.Entry(frame)
        self.passwd = ttk.Entry(frame, show="•")
        self.user.grid(row=0, column=1, pady=6, padx=6)
        self.passwd.grid(row=1, column=1, pady=6, padx=6)

        companies = getattr(master, 'companies', []) or []
        self.company_cb = ttk.Combobox(frame, values=companies, state='readonly')
        if companies:
            self.company_cb.current(0)
        self.company_cb.grid(row=2, column=1, pady=6, padx=6)

        # attempts label
        self.attempts_label = None
        if attempts_left is not None:
            try:
                self.attempts_label = ttk.Label(frame, text=f"Pozostałe próby: {attempts_left}")
                self.attempts_label.grid(row=3, column=0, columnspan=2, pady=(4, 8))
            except Exception:
                self.attempts_label = None

        self.action_var = tk.StringVar(value='')
        self.submitted_username = ''
        self.submitted_password = ''
        self.selected_company = ''

        login_btn = create_btn(self, frame, text="Zaloguj", command=self.ok, bootstyle='warning')
        login_btn.grid(row=4, column=0, columnspan=2, pady=10)

        # ensure dialog is visible and has focus, and bind Enter to submit
        try:
            self.deiconify()
            self.lift()
        except Exception:
            pass
        try:
            self.user.focus_set()
        except Exception:
            pass
        try:
            self.bind('<Return>', lambda e: self.ok())
            self.bind('<KP_Enter>', lambda e: self.ok())
        except Exception:
            pass

    def ok(self):
        self.submitted_username = self.user.get()
        self.submitted_password = self.passwd.get()
        try:
            self.selected_company = self.company_cb.get()
        except Exception:
            self.selected_company = ''
        try:
            self.action_var.set('ok')
        except Exception:
            pass

    def cancel(self):
        try:
            self.action_var.set('cancel')
        except Exception:
            pass
        try:
            self.destroy()
        except Exception:
            pass

    def update_attempts(self, attempts_left: int):
        try:
            if self.attempts_label:
                self.attempts_label.config(text=f"Pozostałe próby: {attempts_left}")
        except Exception:
            pass

# Dialogy pomocnicze (ItemDialog, AccountDialog, RMKAccountDialog, UserDialog, CompanyDialog)
class ItemDialog(tk.Toplevel):
    def __init__(self, master, item: Optional[RMKItem] = None):
        super().__init__(master)
        self.title("Pozycja RMK" if not item else f"Edytuj RMK #{item.id}")
        self.transient(master)
        try:
            self.grab_set()
        except Exception:
            pass
        self.resizable(False, False)
        self.item = item

        frame = ttk.Frame(self, padding=12)
        frame.pack(fill=tk.BOTH, expand=True)

        labels = ["Opis", "Data start (YYYY-MM-DD)", "Data końca (YYYY-MM-DD)", "Liczba miesięcy", "Kwota", "Kategoria", "Konto kosztowe", "Konto RMK", "Nr faktury", "Kontrahent", "Uwagi"]
        labels_with_company = ["Firma"] + labels
        for i, txt in enumerate(labels_with_company):
            ttk.Label(frame, text=txt + ":").grid(row=i, column=0, sticky="e", padx=6, pady=4)

        # widgets
        # If the logged-in user is admin allow selecting any company,
        # otherwise fix the company to the currently selected company at login
        try:
            is_admin = bool(getattr(master, 'current_user_admin', False))
        except Exception:
            is_admin = False
        if is_admin:
            print(f"FIRMA Debug dialog - admin widzi firmy: {master.companies}")
            self.cbx_firma = ttk.Combobox(frame, values=master.companies, state='readonly')
            if master.companies:
                self.cbx_firma.current(0)
        else:
            cur = getattr(master, 'current_company', master.companies[0] if master.companies else '')
            print(f"FIRMA Debug dialog - user widzi firmę: {cur}")
            # show only current company and disable changes for non-admins
            self.cbx_firma = ttk.Combobox(frame, values=[cur], state='disabled')
            if cur:
                try:
                    self.cbx_firma.set(cur)
                except Exception:
                    pass
        self.ent_opis = ttk.Entry(frame, width=40)
        self.ent_data = ttk.Entry(frame)
        self.ent_data_koniec = ttk.Entry(frame)
        self.ent_mies = ttk.Entry(frame)
        self.ent_kwota = ttk.Entry(frame)
        self.cbx_kat = ttk.Combobox(frame, values=master.categories, state='readonly')
        
        # Inicjalnie puste konta - wypełnią się po wyborze firmy
        self.cbx_konto = ttk.Combobox(frame, values=[], state='readonly')
        self.cbx_konto_rmk = ttk.Combobox(frame, values=[], state='readonly')
        
        self.ent_fv = ttk.Entry(frame)
        self.ent_kontr = ttk.Entry(frame)
        self.ent_uwagi = ttk.Entry(frame, width=40)
        
        # Funkcja do aktualizacji kont po wyborze firmy
        def on_company_change(event):
            selected_company = self.cbx_firma.get()
            if selected_company:
                # Zaktualizuj konta kosztowe
                accounts = master.get_accounts_for_company(selected_company)
                self.cbx_konto['values'] = [a['konto'] for a in accounts]
                if accounts:
                    self.cbx_konto.set(accounts[0]['konto'])
                else:
                    self.cbx_konto.set("")
                
                # Zaktualizuj konta RMK
                rmk_accounts = master.get_rmk_accounts_for_company(selected_company)
                self.cbx_konto_rmk['values'] = [r['konto'] for r in rmk_accounts]
                if rmk_accounts:
                    self.cbx_konto_rmk.set(rmk_accounts[0]['konto'])
                else:
                    self.cbx_konto_rmk.set("")
        
        # Podłącz event do zmiany firmy
        if is_admin:
            self.cbx_firma.bind('<<ComboboxSelected>>', on_company_change)

        widgets = [self.cbx_firma, self.ent_opis, self.ent_data, self.ent_data_koniec, self.ent_mies, self.ent_kwota, self.cbx_kat, self.cbx_konto, self.cbx_konto_rmk, self.ent_fv, self.ent_kontr, self.ent_uwagi]
        for i, w in enumerate(widgets):
            w.grid(row=i, column=1, sticky="w", padx=6, pady=4)

        if item:
            try:
                self.cbx_firma.set(item.firma)
                # Najpierw ustaw firmę, potem uruchom aktualizację kont
                if is_admin:
                    on_company_change(None)  # Trigger aktualizacji kont
            except Exception:
                pass
            self.ent_opis.insert(0, item.opis)
            self.ent_data.insert(0, item.data_start.isoformat())
            self.ent_data_koniec.insert(0, item.data_koniec.isoformat() if item.data_koniec else "")
            self.ent_mies.insert(0, str(item.liczba_mies))
            self.ent_kwota.insert(0, str(item.kwota))
            self.cbx_kat.set(item.kategoria)
            self.cbx_konto.set(item.konto_kosztowe)
            self.cbx_konto_rmk.set(item.konto_rmk)
            self.ent_fv.insert(0, item.numer_faktury)
            self.ent_kontr.insert(0, item.kontrahent)
            self.ent_uwagi.insert(0, getattr(item, 'uwagi', '') or '')  # Obsługa starych danych
        else:
            self.ent_data.insert(0, date.today().isoformat())
            # Domyślnie ustaw datę końca na jeden miesiąc później
            default_end = month_add(date.today(), 1)
            self.ent_data_koniec.insert(0, default_end.isoformat())
            self.cbx_kat.set(master.categories[0] if master.categories else "")
            
            # Dla nowych pozycji: ustaw domyślną firmę i załaduj jej konta
            if is_admin and master.companies:
                default_company = master.companies[0]
                self.cbx_firma.set(default_company)
                on_company_change(None)  # Trigger aktualizacji kont
            elif not is_admin:
                # Dla nie-adminów: użyj aktualnej firmy
                current_company = getattr(master, 'current_company', '')
                if current_company:
                    accounts = master.get_accounts_for_company(current_company)
                    rmk_accounts = master.get_rmk_accounts_for_company(current_company)
                    self.cbx_konto['values'] = [a['konto'] for a in accounts]
                    self.cbx_konto_rmk['values'] = [r['konto'] for r in rmk_accounts]
                    if accounts:
                        self.cbx_konto.set(accounts[0]['konto'])
                    if rmk_accounts:
                        self.cbx_konto_rmk.set(rmk_accounts[0]['konto'])
            
            self.ent_mies.insert(0, "12")
            self.ent_kwota.insert(0, "0.00")

        btnf = ttk.Frame(frame)
        btnf.grid(row=len(labels_with_company), column=0, columnspan=2, pady=(8, 0))
        ttk.Button(btnf, text="Anuluj", command=self.cancel).pack(side=tk.RIGHT, padx=6)
        create_btn(self, btnf, text="Zapisz", command=self.ok, bootstyle='warning').pack(side=tk.RIGHT, padx=6)
        self.result = None
        self.wait_window(self)

    def ok(self):
        try:
            firma = self.cbx_firma.get().strip()
            if not firma:
                firma = (self.master.current_company if hasattr(self.master, 'current_company') else '')
            opis = self.ent_opis.get().strip()
            data_start = datetime.strptime(self.ent_data.get().strip(), "%Y-%m-%d").date()
            
            # Spróbuj pobrać datę końca z pola, jeśli puste oblicz automatycznie
            data_koniec_str = self.ent_data_koniec.get().strip()
            if data_koniec_str:
                data_koniec = datetime.strptime(data_koniec_str, "%Y-%m-%d").date()
            else:
                liczba_mies = int(self.ent_mies.get().strip())
                if liczba_mies <= 0:
                    raise ValueError("Liczba miesięcy musi być > 0")
                data_koniec = month_add(data_start, liczba_mies - 1)
            
            # Oblicz liczbę miesięcy z dat jeśli nie podana
            liczba_mies_str = self.ent_mies.get().strip()
            if liczba_mies_str:
                liczba_mies = int(liczba_mies_str)
                if liczba_mies <= 0:
                    raise ValueError("Liczba miesięcy musi być > 0")
            else:
                # Oblicz automatycznie z różnicy dat
                diff = data_koniec - data_start
                liczba_mies = max(1, (diff.days // 30) + 1)  # Aproximacja miesięcy
            
            kwota = float(self.ent_kwota.get().strip())
            kategoria = self.cbx_kat.get().strip()
            konto_kosztowe = self.cbx_konto.get().strip()
            konto_rmk = self.cbx_konto_rmk.get().strip()
            numer_faktury = self.ent_fv.get().strip()
            kontrahent = self.ent_kontr.get().strip()
            uwagi = self.ent_uwagi.get().strip()
        except Exception as e:
            messagebox.showerror("Błąd", f"Niepoprawne dane: {e}")
            return
        self.result = {
            "firma": firma,
            "opis": opis,
            "data_start": data_start,
            "liczba_mies": liczba_mies,
            "kwota": kwota,
            "kategoria": kategoria,
            "konto_kosztowe": konto_kosztowe,
            "konto_rmk": konto_rmk,
            "numer_faktury": numer_faktury,
            "kontrahent": kontrahent,
            "uwagi": uwagi,
            "data_koniec": data_koniec
        }
        self.destroy()

    def cancel(self):
        self.result = None
        self.destroy()

class AccountDialog(simpledialog.Dialog):
    def __init__(self, master, title="Konto księgowe", konto:str="", opis:str=""):
        self.init_konto = konto
        self.init_opis = opis
        super().__init__(master, title=title)
    def body(self, master):
        ttk.Label(master, text="Konto:").grid(row=0, column=0, padx=6, pady=6)
        ttk.Label(master, text="Opis:").grid(row=1, column=0, padx=6, pady=6)
        self.ent_k = ttk.Entry(master, width=30)
        self.ent_o = ttk.Entry(master, width=40)
        self.ent_k.grid(row=0, column=1, padx=6, pady=6)
        self.ent_o.grid(row=1, column=1, padx=6, pady=6)
        self.ent_k.insert(0, self.init_konto)
        self.ent_o.insert(0, self.init_opis)
        return self.ent_k
    def apply(self):
        k = self.ent_k.get().strip()
        o = self.ent_o.get().strip()
        self.result = (k, o)

class RMKAccountDialog(simpledialog.Dialog):
    def __init__(self, master, title="Konto RMK", konto:str="", opis:str=""):
        self.init_konto = konto
        self.init_opis = opis
        super().__init__(master, title=title)
    def body(self, master):
        ttk.Label(master, text="Konto RMK:").grid(row=0, column=0, padx=6, pady=6)
        ttk.Label(master, text="Opis:").grid(row=1, column=0, padx=6, pady=6)
        self.ent_k = ttk.Entry(master, width=30)
        self.ent_o = ttk.Entry(master, width=40)
        self.ent_k.grid(row=0, column=1, padx=6, pady=6)
        self.ent_o.grid(row=1, column=1, padx=6, pady=6)
        self.ent_k.insert(0, self.init_konto)
        self.ent_o.insert(0, self.init_opis)
        return self.ent_k
    def apply(self):
        k = self.ent_k.get().strip()
        o = self.ent_o.get().strip()
        self.result = (k, o)

class UserDialog(tk.Toplevel):
    def __init__(self, master, username: Optional[str] = None, data: Optional[dict] = None):
        super().__init__(master)
        self.title("Użytkownik" if not username else f"Edytuj {username}")
        self.transient(master)
        self.grab_set()
        frame = ttk.Frame(self, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frame, text="Nazwa:").grid(row=0, column=0, sticky='e', pady=4, padx=4)
        ttk.Label(frame, text="Hasło:").grid(row=1, column=0, sticky='e', pady=4, padx=4)
        ttk.Label(frame, text="Admin:").grid(row=2, column=0, sticky='e', pady=4, padx=4)
        ttk.Label(frame, text="Firmy (Ctrl+klik/Shift+klik):").grid(row=3, column=0, sticky='ne', pady=4, padx=4)
        self.ent_name = ttk.Entry(frame)
        self.ent_pwd = ttk.Entry(frame)
        self.var_admin = tk.BooleanVar()
        self.chk_admin = ttk.Checkbutton(frame, variable=self.var_admin)
        self.lb_comp = tk.Listbox(frame, selectmode=tk.MULTIPLE, height=6, exportselection=False)
        for c in master.companies:
            self.lb_comp.insert(tk.END, c)
        self.ent_name.grid(row=0, column=1, padx=6, pady=4, sticky='we')
        self.ent_pwd.grid(row=1, column=1, padx=6, pady=4, sticky='we')
        self.chk_admin.grid(row=2, column=1, padx=6, pady=4, sticky='w')
        self.lb_comp.grid(row=3, column=1, padx=6, pady=4, sticky='we')
        if username and data:
            self.ent_name.insert(0, username)
            self.ent_name.config(state='disabled')
            self.ent_pwd.insert(0, data.get('password', ''))
            self.var_admin.set(data.get('is_admin', False))
            user_comps = data.get('companies', [])
            for i, c in enumerate(master.companies):
                if c in user_comps:
                    self.lb_comp.selection_set(i)
        btnf = ttk.Frame(frame)
        btnf.grid(row=4, column=0, columnspan=2, pady=8)
        ttk.Button(btnf, text="Anuluj", command=self.cancel).pack(side=tk.RIGHT, padx=6)
        create_btn(self, btnf, text="Zapisz", command=self.ok, bootstyle='warning').pack(side=tk.RIGHT, padx=6)
        self.result = None
        self.wait_window(self)

    def ok(self):
        name = self.ent_name.get().strip()
        pwd = self.ent_pwd.get().strip()
        is_admin = bool(self.var_admin.get())
        sel = [self.lb_comp.get(i) for i in self.lb_comp.curselection()]
        if not name or not pwd:
            messagebox.showerror("Błąd", "Nazwa i hasło są wymagane.")
            return
        self.result = {"name": name, "password": pwd, "is_admin": is_admin, "companies": sel}
        self.destroy()

    def cancel(self):
        self.result = None
        self.destroy()

class CompanyDialog(simpledialog.Dialog):
    def __init__(self, master, title="Firma", name: str = ""):
        self.init_name = name
        super().__init__(master, title=title)
    def body(self, master):
        ttk.Label(master, text="Nazwa firmy:").grid(row=0, column=0, padx=6, pady=6)
        self.ent = ttk.Entry(master, width=40)
        self.ent.grid(row=0, column=1, padx=6, pady=6)
        self.ent.insert(0, self.init_name)
        return self.ent
    def apply(self):
        self.result = self.ent.get().strip()

class RMKApp(tk.Tk):
    def __init__(self):
        super().__init__()
        
        # Poprawka dla Windows - ustaw kodowanie konsoli
        if os.name == 'nt':
            try:
                import locale
                locale.setlocale(locale.LC_ALL, '')
            except Exception:
                pass
        
        # Konfiguruj style TTK przed utworzeniem interfejsu
        configure_ttk_styles(self)
        
        self.withdraw()
        self.title(APP_NAME)
        self.geometry("1200x760")
        self.configure(bg=BRAND_COLOR_LIGHT)  # Używaj lekko szarego tła
        
        # Ustaw ikonę okna jeśli istnieje
        try:
            icon_path = resource_path('logo/RMK insG.png')
            if os.path.exists(icon_path):
                # Konwertuj PNG na PhotoImage dla Tkinter
                icon_img = tk.PhotoImage(file=icon_path)
                self.iconphoto(False, icon_img)
        except Exception:
            pass  # Ignoruj błędy z ikoną
        
        self._setup_style()
        self._init_data()
        # load persisted state (if any)
        self._load_state()

        print(">>> Debug main - tworzę splash screen")
        Splash(self)
        print(">>> Debug main - czekam na zamknięcie splash")
        self.wait_window(self.winfo_children()[-1])
        print(">>> Debug main - splash zamknięty")

        # Keep the main window hidden until successful login
        max_attempts = 3
        attempt = 0
        logged_in = False
        # create one dialog and update it dynamically to avoid flicker
        attempts_left = max_attempts - attempt
        login = LoginDialog(self, attempts_left=attempts_left)
        # keep dialog visible and listen to its action_var
        while attempt < max_attempts and not logged_in:
            # wait until dialog signals an action (ok/cancel)
            # use variable trace-like polling since wait_variable on external var isn't safe across Toplevels
            self.wait_variable(login.action_var)
            action = login.action_var.get()
            if action == 'cancel':
                try:
                    login.destroy()
                except Exception:
                    pass
                try:
                    self.destroy()
                except Exception:
                    pass
                sys.exit(0)
            elif action == 'ok':
                username = login.submitted_username
                password = login.submitted_password
                if username in self.users and self.users[username]['password'] == password:
                    logged_in = True
                    self.current_user = username
                    self.current_user_admin = self.users[username].get('is_admin', False)
                    try:
                        login.destroy()
                    except Exception:
                        pass
                    # set current company from login dialog (if provided)
                    try:
                        selc = getattr(login, 'selected_company', None)
                        if selc:
                            self.current_company = selc
                        else:
                            self.current_company = self.companies[0] if self.companies else ''
                    except Exception:
                        self.current_company = self.companies[0] if self.companies else ''
                    break
                else:
                    attempt += 1
                    attempts_left = max_attempts - attempt
                    messagebox.showerror(APP_NAME, "Nieprawidłowy login lub hasło.")
                    # update attempts label on the dialog
                    try:
                        login.update_attempts(attempts_left)
                    except Exception:
                        pass
                    # reset action_var to wait for next submit
                    login.action_var.set('')
                    # continue loop and wait for next user submission

        if not logged_in:
            # exceeded attempts — exit
            try:
                self.destroy()
            except Exception:
                pass
            sys.exit(0)

        # successful login -> show main UI
        self.deiconify()
        # set window title to include selected company (if any)
        try:
            comp = getattr(self, 'current_company', '')
            if comp:
                self.title(f"{APP_NAME} - {comp}")
        except Exception:
            pass
        try:
            self._build_gui()
        except Exception:
            tb = traceback.format_exc()
            try:
                with open('/workspaces/RMK/error.log', 'a', encoding='utf-8') as f:
                    f.write(f"\n[{datetime.now().isoformat()}] Exception in _build_gui():\n")
                    f.write(tb)
                    f.write("\n")
            except Exception:
                pass
            try:
                messagebox.showerror(APP_NAME, "Błąd podczas budowy interfejsu. Sprawdź /workspaces/RMK/error.log dla szczegółów.")
            except Exception:
                pass
            sys.exit(1)
        self.update_status(f"Zalogowano: {self.current_user}{' (Admin)' if self.current_user_admin else ''}")

    def _init_data(self):
        self.users = {
            'admin': {'password': 'admin', 'is_admin': True, 'companies': ['IntegritasAD']},
            'user': {'password': 'user', 'is_admin': False, 'companies': ['TestFirma']}
        }
        self.companies: List[str] = ["IntegritasAD", "TestFirma"]
        self.rmk_items: List[RMKItem] = []
        self.categories = ["Ubezpieczenia", "Licencje", "Najem", "Subskrypcje"]
        
        # Konta kosztowe per firma
        self.accounts_by_company = {
            "IntegritasAD": [
                {"konto": "401-1", "opis": "Koszty podstawowe IntegritasAD"},
                {"konto": "640", "opis": "Rozliczenia międzyokresowe IntegritasAD"}
            ],
            "TestFirma": [
                {"konto": "401-2", "opis": "Koszty podstawowe TestFirma"},
                {"konto": "640-T", "opis": "Rozliczenia międzyokresowe TestFirma"}
            ]
        }
        
        # Konta RMK per firma  
        self.rmk_accounts_by_company = {
            "IntegritasAD": [
                {"konto": "640", "opis": "Rozliczenia międzyokresowe IntegritasAD"},
                {"konto": "640-1", "opis": "RMK - amortyzacja 1 IntegritasAD"}
            ],
            "TestFirma": [
                {"konto": "640-T", "opis": "Rozliczenia międzyokresowe TestFirma"},
                {"konto": "640-T1", "opis": "RMK - amortyzacja 1 TestFirma"}
            ]
        }
        
        # Zachowaj stare formaty dla kompatybilności wstecznej
        self.accounts = []
        self.rmk_accounts = []

    def get_accounts_for_company(self, company: str) -> List[Dict]:
        """Zwraca konta kosztowe dla konkretnej firmy"""
        if not company:
            return []
        return self.accounts_by_company.get(company, [])
    
    def get_rmk_accounts_for_company(self, company: str) -> List[Dict]:
        """Zwraca konta RMK dla konkretnej firmy"""
        if not company:
            return []
        return self.rmk_accounts_by_company.get(company, [])
    
    def add_account_to_company(self, company: str, konto: str, opis: str):
        """Dodaje konto kosztowe do konkretnej firmy"""
        if company not in self.accounts_by_company:
            self.accounts_by_company[company] = []
        
        # Sprawdź czy konto już istnieje
        if not any(a['konto'] == konto for a in self.accounts_by_company[company]):
            self.accounts_by_company[company].append({"konto": konto, "opis": opis})
    
    def add_rmk_account_to_company(self, company: str, konto: str, opis: str):
        """Dodaje konto RMK do konkretnej firmy"""
        if company not in self.rmk_accounts_by_company:
            self.rmk_accounts_by_company[company] = []
        
        # Sprawdź czy konto już istnieje
        if not any(r['konto'] == konto for r in self.rmk_accounts_by_company[company]):
            self.rmk_accounts_by_company[company].append({"konto": konto, "opis": opis})

    # ---- persistence ----
    def _state_file(self):
        # Sprawdź czy jesteśmy w PyInstaller exe
        if getattr(sys, 'frozen', False):
            # W exe - zapisuj w folderze użytkownika
            import tempfile
            user_data_dir = os.path.join(os.path.expanduser("~"), "RMK_insGT")
            os.makedirs(user_data_dir, exist_ok=True)
            data_path = os.path.join(user_data_dir, 'data.json')
            print(f"📂 EXE używa pliku danych: {data_path}")
            return data_path
        else:
            # W rozwoju - zapisuj w workspace
            dev_path = '/workspaces/RMK/data.json'
            print(f"📂 PY używa pliku danych: {dev_path}")
            return dev_path

    def _load_state(self):
        path = self._state_file()
        print(f"🔄 Ładuję dane z: {path}")
        try:
            if os.path.exists(path):
                with open(path, 'r', encoding='utf-8') as f:
                    obj = json.load(f)
                self.users = obj.get('users', self.users)
                self.companies = obj.get('companies', self.companies)
                self.categories = obj.get('categories', self.categories)
                self.accounts = obj.get('accounts', self.accounts)
                self.rmk_accounts = obj.get('rmk_accounts', self.rmk_accounts)
                
                # Załaduj nowe struktury per firma
                self.accounts_by_company = obj.get('accounts_by_company', getattr(self, 'accounts_by_company', {}))
                self.rmk_accounts_by_company = obj.get('rmk_accounts_by_company', getattr(self, 'rmk_accounts_by_company', {}))
                
                self.rmk_items = [RMKItem.from_dict(d) for d in obj.get('rmk_items', [])]
                
                # Zapewnij kompatybilność wsteczną - dodaj puste uwagi do starych pozycji
                for item in self.rmk_items:
                    if not hasattr(item, 'uwagi'):
                        item.uwagi = ""
                
                self.view_state = obj.get('view_state', {})
                
                # Debug: sprawdź załadowane firmy
                print(f"FIRMA Debug companies - załadowano z pliku: {self.companies}")
                print(f"OK Załadowano dane - firm: {len(self.companies)}, użytkowników: {len(self.users)}")
            else:
                print(f"UWAGA Plik danych nie istnieje: {path}")
        except Exception as e:
            print(f"BŁĄD Błąd ładowania danych: {e}")
            print(f"FIRMA Debug companies - używam domyślnych: {self.companies}")
            # Jeśli nie można załadować, utwórz domyślne dane

    def _save_state(self):
        path = self._state_file()
        obj = {
            'users': self.users,
            'companies': self.companies,
            'categories': self.categories,
            'accounts': self.accounts,  # Zachowaj dla kompatybilności
            'rmk_accounts': self.rmk_accounts,  # Zachowaj dla kompatybilności
            'accounts_by_company': getattr(self, 'accounts_by_company', {}),  # Nowa struktura
            'rmk_accounts_by_company': getattr(self, 'rmk_accounts_by_company', {}),  # Nowa struktura
            'rmk_items': [it.to_dict() for it in self.rmk_items],
            'view_state': getattr(self, 'view_state', {})
        }
        try:
            os.makedirs(os.path.dirname(path), exist_ok=True)
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(obj, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print('Error saving state:', e)

    def create_backup(self):
        """Tworzy kopię zapasową bazy danych"""
        try:
            import shutil
            from datetime import datetime
            
            source_path = self._state_file()
            if not os.path.exists(source_path):
                messagebox.showwarning(APP_NAME, "Brak pliku danych do backupu.")
                return
            
            # Utwórz folder backupów
            backup_dir = os.path.join(os.path.dirname(source_path), "backups")
            os.makedirs(backup_dir, exist_ok=True)
            
            # Nazwa pliku z datą i czasem
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_filename = f"data_backup_{timestamp}.json"
            backup_path = os.path.join(backup_dir, backup_filename)
            
            # Skopiuj plik
            shutil.copy2(source_path, backup_path)
            
            messagebox.showinfo(APP_NAME, f"Backup utworzony:\n{backup_path}")
            
        except Exception as e:
            messagebox.showerror(APP_NAME, f"Błąd podczas tworzenia backupu:\n{e}")

    def restore_backup(self):
        """Przywraca kopię zapasową bazy danych"""
        try:
            backup_dir = os.path.join(os.path.dirname(self._state_file()), "backups")
            if not os.path.exists(backup_dir):
                messagebox.showwarning(APP_NAME, "Brak foldera z backupami.")
                return
            
            backup_file = filedialog.askopenfilename(
                title="Wybierz plik backupu do przywrócenia",
                initialdir=backup_dir,
                filetypes=[('JSON files', '*.json'), ('All files', '*.*')]
            )
            
            if not backup_file:
                return
            
            if messagebox.askyesno(APP_NAME, "Czy na pewno przywrócić backup?\nAktualne dane zostaną zastąpione!"):
                import shutil
                shutil.copy2(backup_file, self._state_file())
                messagebox.showinfo(APP_NAME, "Backup przywrócony. Uruchom aplikację ponownie.")
                self.quit()
                
        except Exception as e:
            messagebox.showerror(APP_NAME, f"Błąd podczas przywracania backupu:\n{e}")

    def _create_menu(self):
        """Tworzy menu aplikacji"""
        menubar = tk.Menu(self)
        self.config(menu=menubar)
        
        # Menu Plik
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Plik", menu=file_menu)
        file_menu.add_command(label="Utwórz backup", command=self.create_backup)
        file_menu.add_command(label="Przywróć backup", command=self.restore_backup)
        file_menu.add_separator()
        file_menu.add_command(label="Wyjście", command=self.quit)
        
        # Menu Import/Export
        import_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Import/Export", menu=import_menu)
        import_menu.add_command(label="Import Excel", command=self.import_excel)
        # Tutaj można dodać eksport do Excel w przyszłości
        
        # Menu Pomoc
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Pomoc", menu=help_menu)
        help_menu.add_command(label="O programie", command=self.show_about)

    def show_about(self):
        """Wyświetla informacje o programie"""
        about_text = f"""
{APP_NAME}

Aplikacja do zarządzania rozliczeniami 
międzyokresowymi kosztów

Wersja: 0.22.0

Funkcjonalności:
• Zarządzanie pozycjami RMK per firma
• Harmonogramy z precyzyjnym dzieleniem według dni
• Import z Excel z mapowaniem kolumn
• Konta kosztowe i RMK per firma
• Generowanie raportów PDF
• System backupów
• Obsługa polskich znaków

© 2024-2025
        """
        messagebox.showinfo("O programie", about_text.strip())

    def _setup_style(self):
        # Prefer ttkbootstrap Style if available for modern look
        try:
            if TTKBOOTSTRAP_AVAILABLE and TBStyle is not None:
                # initialize ttkbootstrap style with a clean theme
                self.tb_style = TBStyle(theme='litera')
                # configure a custom accent via bootstyle if needed
                # ttkbootstrap uses 'success', 'primary', 'warning' etc. via 'bootstyle' arg
                return
        except Exception:
            pass
        # fallback to plain ttk styling
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('Accent.TButton', background=BRAND_COLOR_YELLOW, foreground='black', relief='flat')
        style.map('Accent.TButton', background=[('active', BRAND_COLOR_ACCENT)])

    def _btn(self, parent, **kwargs):
        """Helper tworzy ttk.Button z bootstyle gdy ttkbootstrap dostępne,
        lub z mapowanym style w przeciwnym przypadku.
        Użycie: self._btn(frame, text='OK', command=..., bootstyle='warning').pack(...)
        """
        # copy kwargs to avoid mutation
        k = dict(kwargs)
        # extract explicit style hints
        bootstyle = k.pop('bootstyle', None)
        style = k.pop('style', None)
        if TTKBOOTSTRAP_AVAILABLE and bootstyle:
            # pass bootstyle to ttk.Button (ttkbootstrap injects support)
            return ttk.Button(parent, bootstyle=bootstyle, **k)
        # fallback: mapuj bootstyle na nasze style
        if not style and bootstyle:
            style_map = {
                'success': 'Success.TButton',
                'warning': 'Warning.TButton', 
                'danger': 'Danger.TButton',
                'primary': 'Info.TButton',
                'info': 'Info.TButton'
            }
            style = style_map.get(bootstyle, 'TButton')
        return ttk.Button(parent, style=(style or 'TButton'), **k)

    def _add_logo_header(self):
        """Dodaj logo w lewym górnym rogu aplikacji"""
        try:
            print("IMG Dodaję logo do głównego okna...")
            
            # Znajdź logo sprawdzając różne lokalizacje
            base_dir = resource_path("")
            logo_path = None
            
            # Lista miejsc do sprawdzenia
            search_locations = []
            
            # 1. Folder logo/ w aplikacji
            try:
                logo_dir = os.path.join(base_dir, 'logo')
                if os.path.isdir(logo_dir):
                    search_locations.append(logo_dir)
            except:
                pass
            
            # 2. Główny katalog aplikacji
            search_locations.append(base_dir)
            
            # 3. Folder domyślny użytkownika
            try:
                home_dir = os.path.expanduser("~")
                search_locations.append(home_dir)
                print(f"🏠 Dodano folder domyślny: {home_dir}")
            except:
                pass
            
            # 4. Pulpit użytkownika (Windows)
            try:
                if os.name == 'nt':
                    desktop_dir = os.path.join(os.path.expanduser("~"), "Desktop")
                    if os.path.isdir(desktop_dir):
                        search_locations.append(desktop_dir)
                        print(f"MONITOR Dodano pulpit: {desktop_dir}")
            except:
                pass
            
            # 5. Dokumenty użytkownika
            try:
                documents_dir = os.path.join(os.path.expanduser("~"), "Documents")
                if os.path.isdir(documents_dir):
                    search_locations.append(documents_dir)
                    print(f"📁 Dodano dokumenty: {documents_dir}")
            except:
                pass
            
            # Nazwy plików logo do sprawdzenia
            logo_filenames = [
                'RMK_insGT_logo.png', 'RMK insG.png',  # Konkretne nazwy z aplikacji
                'logo.png', 'logo.jpg', 'logo.jpeg',   # Ogólne nazwy
                'rmk.png', 'rmk.jpg', 'rmk.jpeg',      # RMK nazwy
                'company_logo.png', 'firma_logo.png'    # Firmowe nazwy
            ]
            
            print(f"🔍 Sprawdzam {len(search_locations)} lokalizacji...")
            print(f"📋 Szukam plików: {logo_filenames}")
            
            # Przeszukaj wszystkie lokalizacje
            for location in search_locations:
                print(f"📂 Sprawdzam: {location}")
                try:
                    if os.path.isdir(location):
                        # Pokaż co jest w folderze
                        files_in_dir = os.listdir(location)
                        print(f"   FILES Pliki w folderze: {[f for f in files_in_dir if f.lower().endswith(('.png', '.jpg', '.jpeg'))]}")
                        
                        for filename in logo_filenames:
                            potential_path = os.path.join(location, filename)
                            if os.path.exists(potential_path):
                                logo_path = potential_path
                                print(f"OK ZNALEZIONO LOGO: {logo_path}")
                                break
                        
                        if logo_path:
                            break
                            
                        # Dodatkowo sprawdź wszystkie pliki PNG/JPG w folderze
                        try:
                            for file in os.listdir(location):
                                if file.lower().endswith(('.png', '.jpg', '.jpeg')) and 'logo' in file.lower():
                                    potential_path = os.path.join(location, file)
                                    if os.path.exists(potential_path):
                                        logo_path = potential_path
                                        print(f"OK ZNALEZIONO LOGO (wzorzec): {logo_path}")
                                        break
                            if logo_path:
                                break
                        except:
                            pass
                            
                except Exception as e:
                    print(f"UWAGA Błąd sprawdzania {location}: {e}")
            
            if logo_path and os.path.exists(logo_path):
                try:
                    # Stwórz header frame - jasny szary nagłówek z logo (mniejszy)
                    header_frame = tk.Frame(self, bg=BRAND_COLOR_HEADER, height=50)
                    header_frame.pack(fill=tk.X, pady=(0, 4))
                    header_frame.pack_propagate(False)  # Zachowaj stałą wysokość
                    
                    # Załaduj i przeskaluj logo
                    logo_img = tk.PhotoImage(file=logo_path)
                    
                    # Przeskaluj do maksymalnie 40px wysokości
                    original_width = logo_img.width()
                    original_height = logo_img.height()
                    max_height = 40
                    
                    if original_height > max_height:
                        scale_factor = max_height / original_height
                        new_width = int(original_width * scale_factor)
                        new_height = max_height
                        
                        # Przeskaluj obraz
                        logo_img = logo_img.subsample(int(1/scale_factor))
                    
                    # Dodaj logo do lewej strony
                    logo_label = tk.Label(header_frame, 
                                        image=logo_img, 
                                        bg=BRAND_COLOR_HEADER,
                                        bd=0)
                    logo_label.image = logo_img  # Zachowaj referencję
                    logo_label.pack(side=tk.LEFT, padx=(8, 4), pady=4)
                    
                    # Dodaj tytuł aplikacji obok logo w stylu SAP
                    title_label = tk.Label(header_frame,
                                         text=f"{COMPANY_NAME} | {APP_NAME}",
                                         font=("Segoe UI", 12, "bold"),
                                         fg=BRAND_COLOR_TEXT,  # Ciemnoszary tekst SAP
                                         bg=BRAND_COLOR_HEADER)
                    title_label.pack(side=tk.LEFT, padx=(4, 0), pady=4, anchor='w')
                    
                    print(f"OK Logo dodane do głównego okna ({original_width}x{original_height}) z: {logo_path}")
                    
                except Exception as e:
                    print(f"UWAGA Błąd ładowania logo do GUI: {e}")
                    # Dodaj placeholder jeśli logo się nie załaduje
                    self._add_text_header()
            else:
                print("UWAGA Nie znaleziono logo w żadnej lokalizacji - dodaję tylko tytuł")
                print(f"🔍 Sprawdzono lokalizacje: {search_locations}")
                self._add_text_header()
                
        except Exception as e:
            print(f"UWAGA Błąd dodawania logo header: {e}")
            # Fallback - tylko tekst
            self._add_text_header()
    
    def _add_text_header(self):
        """Fallback header z samym tekstem"""
        try:
            header_frame = tk.Frame(self, bg=BRAND_COLOR_HEADER, height=50)
            header_frame.pack(fill=tk.X, pady=(0, 4))
            header_frame.pack_propagate(False)
            
            title_label = tk.Label(header_frame,
                                 text=f"{COMPANY_NAME} | {APP_NAME}",
                                 font=("Segoe UI", 14, "bold"),
                                 fg=BRAND_COLOR_TEXT,  # Ciemnoszary tekst SAP
                                 bg=BRAND_COLOR_HEADER)
            title_label.pack(pady=8)
        except Exception as e:
            print(f"UWAGA Błąd text header: {e}")

    def _build_gui(self):
        self.status_var = tk.StringVar()
        
        # Dodaj menu
        self._create_menu()
        
        # Dodaj logo w lewym górnym rogu
        self._add_logo_header()
        
        nb = ttk.Notebook(self)
        nb.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        self._build_tab_lista(nb)
        self._build_tab_harmonogram(nb)
        # dodatkowa zakładka: podsumowanie (agregacje)
        self._build_tab_podsumowanie(nb)
        self._build_tab_rmk_next_year(nb)
        self._build_tab_rmk_by_years(nb)
        self._build_tab_slownik(nb)
        self._build_tab_admin(nb)
        self._build_tab_reports(nb)

        status = ttk.Label(self, textvariable=self.status_var, anchor='w', background=BRAND_COLOR_ACCENT, foreground='white', padding=4)
        status.pack(side=tk.BOTTOM, fill=tk.X)

    def _build_tab_lista(self, nb):
        tab = ttk.Frame(nb)
        tab.configure(style='Light.TFrame')  # Lekko szare tło dla zakładki
        nb.add(tab, text="Lista RMK")
        toolbar = ttk.Frame(tab)
        toolbar.pack(fill=tk.X, padx=8, pady=6)
        create_btn(self, toolbar, text="+ Dodaj", command=self.add_item, bootstyle='success').pack(side=tk.LEFT, padx=4)
        ttk.Button(toolbar, text="Edytuj", command=self.edit_item).pack(side=tk.LEFT, padx=4)
        ttk.Button(toolbar, text="Usuń", command=self.delete_item).pack(side=tk.LEFT, padx=4)
        create_btn(self, toolbar, text="Generuj harmonogram", command=self.generate_harmonogram, bootstyle='primary').pack(side=tk.LEFT, padx=4)
        ttk.Button(toolbar, text="Import Excel", command=self.import_excel).pack(side=tk.LEFT, padx=4)
        # company switch (allow changing current company context)
        ttk.Button(toolbar, text="🔁 Zmień firmę", command=self._change_company).pack(side=tk.RIGHT, padx=4)
        columns = ("id", "opis", "data_start", "data_koniec", "liczba_mies", "kwota", "kategoria", "konto", "konto_rmk", "faktura", "kontrahent", "status", "uwagi")
        
        # Dodaj Frame z ramką wokół tabeli
        # Główna ramka tabeli w stylu SAP - białe tło z szarym obramowaniem
        table_frame = tk.Frame(tab, bd=1, relief='solid', bg=BRAND_COLOR_FRAME, highlightbackground=BRAND_COLOR_BORDER, highlightthickness=1)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=6)
        
        # Utwórz frame dla tabeli ze scrollbarami
        tree_scroll_frame = ttk.Frame(table_frame)
        tree_scroll_frame.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)
        
        self.tree = ttk.Treeview(tree_scroll_frame, columns=columns, show='headings', height=14)
        
        # Wymuś separatory i obramowania
        configure_single_treeview_borders(self.tree)
        
        # Dodaj scrollbary
        v_scrollbar = ttk.Scrollbar(tree_scroll_frame, orient="vertical", command=self.tree.yview)
        h_scrollbar = ttk.Scrollbar(tree_scroll_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Pack scrollbary i treeview
        v_scrollbar.pack(side="right", fill="y")
        h_scrollbar.pack(side="bottom", fill="x")
        self.tree.pack(side="left", fill=tk.BOTH, expand=True)
        for c in columns:
            # Ustaw odpowiednie polskie nagłówki
            if c == "id":
                header_text = "ID"
            elif c == "opis":
                header_text = "Opis"
            elif c == "data_start":
                header_text = "Data start"
            elif c == "data_koniec":
                header_text = "Data koniec"
            elif c == "liczba_mies":
                header_text = "Liczba miesięcy"
            elif c == "kwota":
                header_text = "Kwota"
            elif c == "kategoria":
                header_text = "Kategoria"
            elif c == "konto":
                header_text = "Konto"
            elif c == "konto_rmk":
                header_text = "Konto RMK"
            elif c == "faktura":
                header_text = "Faktura"
            elif c == "kontrahent":
                header_text = "Kontrahent"
            elif c == "status":
                header_text = "Status"
            elif c == "uwagi":
                header_text = "Uwagi"
            else:
                # Dla innych kolumn użyj bezpiecznego formatowania
                header_text = str(c).replace('_', ' ')
                if header_text:
                    header_text = header_text[0].upper() + header_text[1:]
            self.tree.heading(c, text=header_text)
        # configure column widths and alignment (amounts right-aligned)
        self.tree.column('id', width=50, anchor='center')
        self.tree.column('opis', width=300, anchor='w')
        self.tree.column('data_start', width=110, anchor='center')
        self.tree.column('data_koniec', width=110, anchor='center')
        self.tree.column('liczba_mies', width=80, anchor='e')
        self.tree.column('kwota', width=120, anchor='e')
        self.tree.column('kategoria', width=140, anchor='w')
        self.tree.column('konto', width=100, anchor='w')
        self.tree.column('konto_rmk', width=100, anchor='w')
        self.tree.column('faktura', width=110, anchor='w')
        self.tree.column('kontrahent', width=160, anchor='w')
        self.tree.column('status', width=60, anchor='center')
        self.tree.column('uwagi', width=150, anchor='w')
        
        # Dodaj obramowania do tabeli Lista RMK
        configure_single_treeview_borders(self.tree)
        
        self.refresh_rmk_tree()

    def _build_tab_harmonogram(self, nb):
        tab = ttk.Frame(nb)
        tab.configure(style='Light.TFrame')  # Lekko szare tło
        nb.add(tab, text="Harmonogram")
        toolbar = ttk.Frame(tab)
        toolbar.pack(fill=tk.X, padx=8, pady=6)
        # wybór pozycji RMK dla której pokażemy harmonogram
        ttk.Label(toolbar, text="Pozycja RMK:").pack(side=tk.LEFT, padx=4)
        self.harmo_item_cb = ttk.Combobox(toolbar, values=[f"{it.id}: {it.kategoria} | {it.opis} | {it.data_start.strftime('%Y-%m-%d')} - {it.data_koniec.strftime('%Y-%m-%d') if it.data_koniec else 'N/A'}" for it in self.rmk_items], state='readonly', width=80)
        if self.rmk_items:
            self.harmo_item_cb.current(0)
        self.harmo_item_cb.pack(side=tk.LEFT, padx=4)
        create_btn(self, toolbar, text="� Pokaż harmonogram", command=self.show_selected_harmonogram, bootstyle='primary').pack(side=tk.LEFT, padx=4)
        # usuń przycisk filtruj (nieaktywny)
        # USUNIĘTO: przycisk podsumowania - jest w kolejnej zakładce
        
        # Dodaj Frame z ramką wokół tabeli harmonogramu w stylu SAP
        harmo_table_frame = tk.Frame(tab, bd=1, relief='solid', bg=BRAND_COLOR_FRAME, highlightbackground=BRAND_COLOR_BORDER, highlightthickness=1)
        harmo_table_frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=6)
        
        # Utwórz frame dla tabeli harmonogramu ze scrollbarami
        harmo_scroll_frame = ttk.Frame(harmo_table_frame)
        harmo_scroll_frame.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)
        
        # Harmonogram - dynamiczne kolumny będą tworzone przy generowaniu
        self.harmo_tree = ttk.Treeview(harmo_scroll_frame, show='headings', height=16)
        
        # Wymuś separatory i obramowania
        configure_single_treeview_borders(self.harmo_tree)
        
        # Dodaj scrollbary
        harmo_v_scrollbar = ttk.Scrollbar(harmo_scroll_frame, orient="vertical", command=self.harmo_tree.yview)
        harmo_h_scrollbar = ttk.Scrollbar(harmo_scroll_frame, orient="horizontal", command=self.harmo_tree.xview)
        self.harmo_tree.configure(yscrollcommand=harmo_v_scrollbar.set, xscrollcommand=harmo_h_scrollbar.set)
        
        # Pack scrollbary i treeview
        harmo_v_scrollbar.pack(side="right", fill="y")
        harmo_h_scrollbar.pack(side="bottom", fill="x")
        self.harmo_tree.pack(side="left", fill=tk.BOTH, expand=True)
        
        # Dodaj obramowania do tabeli Harmonogram
        configure_single_treeview_borders(self.harmo_tree)

    def _build_tab_podsumowanie(self, nb):
        tab = ttk.Frame(nb)
        tab.configure(style='Light.TFrame')  # Lekko szare tło
        nb.add(tab, text="Podsumowanie")
        toolbar = ttk.Frame(tab)
        toolbar.pack(fill=tk.X, padx=8, pady=6)
        ttk.Label(toolbar, text="Agreguj wg:").pack(side=tk.LEFT, padx=4)
        self.sum_group_cb = ttk.Combobox(toolbar, values=["Kategoria", "Konto kosztowe", "Konto RMK"], state='readonly')
        self.sum_group_cb.current(0)
        self.sum_group_cb.pack(side=tk.LEFT, padx=4)
        ttk.Label(toolbar, text="Zakres: Od (YYYY-MM)").pack(side=tk.LEFT, padx=4)
        self.sum_od = ttk.Entry(toolbar, width=8)
        self.sum_od.pack(side=tk.LEFT, padx=4)
        ttk.Label(toolbar, text="Do (YYYY-MM)").pack(side=tk.LEFT, padx=4)
        self.sum_do = ttk.Entry(toolbar, width=8)
        self.sum_do.pack(side=tk.LEFT, padx=4)
        create_btn(self, toolbar, text="Generuj podsumowanie", command=self.generate_summary, bootstyle='primary').pack(side=tk.LEFT, padx=8)

        self.sum_frame = ttk.Frame(tab)
        self.sum_frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=6)

    def generate_summary(self):
        group = self.sum_group_cb.get()
        od = self.sum_od.get().strip()
        do = self.sum_do.get().strip()
        def parse_ym(s):
            if not s:
                return None
            try:
                return datetime.strptime(s + "-01", "%Y-%m-%d").date()
            except:
                return None
        od_d = parse_ym(od)
        do_d = parse_ym(do)
        if (od and not od_d) or (do and not do_d):
            messagebox.showerror("Błąd", "Niepoprawny format daty (użyj YYYY-MM)")
            return
        # build aggregation key function
        if group == 'Kategoria':
            keyfn = lambda it: it.kategoria
        elif group == 'Konto kosztowe':
            keyfn = lambda it: it.konto_kosztowe
        else:
            keyfn = lambda it: it.konto_rmk

        try:
            is_admin = bool(getattr(self, 'current_user_admin', False))
        except Exception:
            is_admin = False
        cur_company = getattr(self, 'current_company', '')

        agg: Dict[str, Dict[str, float]] = {}
        keys = set()
        for it in self.rmk_items:
            if cur_company and it.firma and it.firma != cur_company:
                continue
            # Użyj ulepszonego algorytmu rozliczania miesięcznego
            monthly_amounts = calculate_monthly_amounts_improved(it.kwota, it.data_start, it.liczba_mies)
            
            for i in range(it.liczba_mies):
                mdate = month_add(it.data_start, i)
                if od_d and mdate < od_d: continue
                if do_d and mdate > do_d: continue
                m = mdate.strftime("%Y-%m")
                k = keyfn(it) or ""
                keys.add(k)
                agg.setdefault(m, {})
                
                # Użyj odpowiedniej kwoty z ulepszonego algorytmu
                amount_to_add = monthly_amounts[i] if i < len(monthly_amounts) else 0.0
                agg[m][k] = agg[m].get(k, 0.0) + amount_to_add

        for w in self.sum_frame.winfo_children():
            w.destroy()
        months = sorted(agg.keys())
        cols = ["Grupa"] + months + ["Razem"]
        tree_container, tree = self._make_scrolled_tree(self.sum_frame, cols)

        for k in sorted(keys):
            row = [k]
            total = 0.0
            for m in months:
                v = agg.get(m, {}).get(k, 0.0)
                total += v
                row.append(thousand_sep(v))
            row.append(thousand_sep(total))
            tree.insert('', 'end', values=row)
        tree_container.pack(fill=tk.BOTH, expand=True)

    def _build_tab_rmk_next_year(self, nb):
        tab = ttk.Frame(nb)
        tab.configure(style='Light.TFrame')  # Lekko szare tło
        nb.add(tab, text="RMK - następny rok")
        toolbar = ttk.Frame(tab)
        toolbar.pack(fill=tk.X, padx=8, pady=6)
        ttk.Label(toolbar, text="Rok: ").pack(side=tk.LEFT, padx=4)
        # prepare years list (current year .. current year+5)
        yrs = [str(date.today().year + i) for i in range(0, 6)]
        self.rmk_year_cb = ttk.Combobox(toolbar, values=yrs, state='readonly', width=6)
        # restore last selected year if available
        last = getattr(self, 'view_state', {}).get('rmk_next_year')
        if last and str(last) in yrs:
            self.rmk_year_cb.set(str(last))
        else:
            self.rmk_year_cb.set(yrs[1])
        self.rmk_year_cb.pack(side=tk.LEFT, padx=4)
        ttk.Label(toolbar, text="Kategoria:").pack(side=tk.LEFT, padx=4)
        self.rmk_year_cat_cb = ttk.Combobox(toolbar, values=["Wszystkie"] + self.categories, state='readonly')
        self.rmk_year_cat_cb.current(0)
        self.rmk_year_cat_cb.pack(side=tk.LEFT, padx=4)
        ttk.Label(toolbar, text="Konto RMK:").pack(side=tk.LEFT, padx=4)
        self.rmk_year_rmk_cb = ttk.Combobox(toolbar, values=["Wszystkie"] + [r['konto'] for r in self.rmk_accounts], state='readonly')
        self.rmk_year_rmk_cb.current(0)
        self.rmk_year_rmk_cb.pack(side=tk.LEFT, padx=4)
        
        # Dodaj opcje grupowania
        ttk.Label(toolbar, text="Grupuj wg:").pack(side=tk.LEFT, padx=(12, 4))
        self.rmk_year_group_var = tk.StringVar(value="kategoria")
        ttk.Radiobutton(toolbar, text="Kategorii", variable=self.rmk_year_group_var, value="kategoria").pack(side=tk.LEFT, padx=2)
        ttk.Radiobutton(toolbar, text="Kont RMK", variable=self.rmk_year_group_var, value="konto_rmk").pack(side=tk.LEFT, padx=2)
        
        self._btn(toolbar, text="Generuj", command=self.generate_rmk_next_year, bootstyle='primary').pack(side=tk.LEFT, padx=8)
        ttk.Button(toolbar, text="Eksportuj PDF", command=self.export_rmk_next_year_pdf).pack(side=tk.LEFT, padx=6)

        self.rmk_year_frame = ttk.Frame(tab)
        self.rmk_year_frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=6)

    def generate_rmk_next_year(self):
        yr = self.rmk_year_cb.get().strip()
        try:
            year = int(yr)
        except Exception:
            messagebox.showerror(APP_NAME, "Niepoprawny rok")
            return
        cat = self.rmk_year_cat_cb.get().strip()
        konto_rmk = self.rmk_year_rmk_cb.get().strip()
        group_by = self.rmk_year_group_var.get()  # "kategoria" lub "konto_rmk"
        cur_company = getattr(self, 'current_company', '')

        # aggregate per month for the chosen year per selected filters
        agg: Dict[str, Dict[str, float]] = {}
        keys = set()
        # months headers for year
        months = [f"{year}-{m:02d}" for m in range(1,13)]

        for it in self.rmk_items:
            # respect current company
            if cur_company and it.firma and it.firma != cur_company:
                continue
            # filter by category and konto_rmk if set
            if cat and cat != "Wszystkie":
                try:
                    if str(it.kategoria).strip().casefold() != cat.casefold():
                        continue
                except Exception:
                    continue
            if konto_rmk and konto_rmk != "Wszystkie":
                if str(it.konto_rmk).strip() != konto_rmk:
                    continue

            # Użyj ulepszonego algorytmu rozliczania miesięcznego
            monthly_amounts = calculate_monthly_amounts_improved(it.kwota, it.data_start, it.liczba_mies)
            
            for i in range(it.liczba_mies):
                mdate = month_add(it.data_start, i)
                if mdate.year != year:
                    continue
                m = mdate.strftime("%Y-%m")
                
                # Wybierz klucz grupowania według ustawienia
                if group_by == "konto_rmk":
                    k = it.konto_rmk or ""
                else:  # domyślnie kategoria
                    k = it.kategoria or ""
                    
                keys.add(k)
                agg.setdefault(k, {})
                
                # Użyj odpowiedniej kwoty z ulepszonego algorytmu
                amount_to_add = monthly_amounts[i] if i < len(monthly_amounts) else 0.0
                agg[k][m] = agg[k].get(m, 0.0) + amount_to_add

        # persist selected year to view_state
        try:
            self.view_state = getattr(self, 'view_state', {})
            self.view_state['rmk_next_year'] = year
        except Exception:
            pass

        # build table with rows = groups (kategorie), columns = months
        for w in self.rmk_year_frame.winfo_children():
            w.destroy()
        cols = ["Kategoria"] + months + ["Razem"]
        tree = ttk.Treeview(self.rmk_year_frame, columns=cols, show='headings')
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=100, anchor='e' if c != 'Kategoria' else 'w')
        tree.pack(fill=tk.BOTH, expand=True)
        
        # Dodaj obramowania do tabeli RMK - następny rok
        configure_single_treeview_borders(tree)

        for k in sorted(keys):
            row = [k]
            total = 0.0
            for m in months:
                v = agg.get(k, {}).get(m, 0.0)
                total += v
                row.append(thousand_sep(v))
            row.append(thousand_sep(total))
            tree.insert('', 'end', values=row)
        # footer: SUMA per month and grand total
        if months:
            footer = ["SUMA"]
            grand = 0.0
            for m in months:
                s = sum(agg.get(k, {}).get(m, 0.0) for k in keys)
                grand += s
                footer.append(thousand_sep(s))
            footer.append(thousand_sep(grand))
            tree.insert('', 'end', values=footer)
        # save view state
        try:
            self._save_state()
        except Exception:
            pass

    def _build_tab_rmk_by_years(self, nb):
        tab = ttk.Frame(nb)
        tab.configure(style='Light.TFrame')  # Lekko szare tło
        nb.add(tab, text="RMK wg lat")
        toolbar = ttk.Frame(tab)
        toolbar.pack(fill=tk.X, padx=8, pady=6)
        ttk.Label(toolbar, text="Rok od:").pack(side=tk.LEFT, padx=4)
        self.rmk_by_year_from = ttk.Combobox(toolbar, values=[str(date.today().year + i) for i in range(-5, 6)], width=6, state='readonly')
        last = getattr(self, 'view_state', {}).get('rmk_by_years', {})
        if last and str(last.get('from')) in self.rmk_by_year_from['values']:
            self.rmk_by_year_from.set(str(last.get('from')))
        else:
            self.rmk_by_year_from.set(str(date.today().year - 1))
        self.rmk_by_year_from.pack(side=tk.LEFT, padx=4)
        ttk.Label(toolbar, text="Rok do:").pack(side=tk.LEFT, padx=4)
        self.rmk_by_year_to = ttk.Combobox(toolbar, values=[str(date.today().year + i) for i in range(-5, 6)], width=6, state='readonly')
        if last and str(last.get('to')) in self.rmk_by_year_to['values']:
            self.rmk_by_year_to.set(str(last.get('to')))
        else:
            self.rmk_by_year_to.set(str(date.today().year + 1))
        self.rmk_by_year_to.pack(side=tk.LEFT, padx=4)
        ttk.Label(toolbar, text="Kategoria:").pack(side=tk.LEFT, padx=4)
        self.rmk_by_year_cat_cb = ttk.Combobox(toolbar, values=["Wszystkie"] + self.categories, state='readonly')
        if last and last.get('category') in (["Wszystkie"] + self.categories):
            try:
                self.rmk_by_year_cat_cb.set(last.get('category'))
            except Exception:
                self.rmk_by_year_cat_cb.current(0)
        else:
            self.rmk_by_year_cat_cb.current(0)
        self.rmk_by_year_cat_cb.pack(side=tk.LEFT, padx=4)
        ttk.Label(toolbar, text="Konto RMK:").pack(side=tk.LEFT, padx=4)
        self.rmk_by_year_rmk_cb = ttk.Combobox(toolbar, values=["Wszystkie"] + [r['konto'] for r in self.rmk_accounts], state='readonly')
        if last and last.get('konto_rmk') in ( ["Wszystkie"] + [r['konto'] for r in self.rmk_accounts] ):
            try:
                self.rmk_by_year_rmk_cb.set(last.get('konto_rmk'))
            except Exception:
                self.rmk_by_year_rmk_cb.current(0)
        else:
            self.rmk_by_year_rmk_cb.current(0)
        self.rmk_by_year_rmk_cb.pack(side=tk.LEFT, padx=4)
        self._btn(toolbar, text="Generuj", command=self.generate_rmk_by_years, bootstyle='primary').pack(side=tk.LEFT, padx=8)

        self.rmk_by_year_frame = ttk.Frame(tab)
        self.rmk_by_year_frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=6)

    def generate_rmk_by_years(self):
        fr = self.rmk_by_year_from.get()
        to = self.rmk_by_year_to.get()
        try:
            y1 = int(fr)
            y2 = int(to)
        except Exception:
            messagebox.showerror(APP_NAME, "Niepoprawny zakres lat")
            return
        if y2 < y1:
            messagebox.showerror(APP_NAME, "Rok do musi być >= Rok od")
            return
        cat = self.rmk_by_year_cat_cb.get().strip()
        konto_rmk = self.rmk_by_year_rmk_cb.get().strip()
        cur_company = getattr(self, 'current_company', '')

        years = [str(y) for y in range(y1, y2 + 1)]
        agg: Dict[str, Dict[str, float]] = {}
        keys = set()

        # persist view state
        try:
            self.view_state = getattr(self, 'view_state', {})
            self.view_state['rmk_by_years'] = {'from': y1, 'to': y2, 'category': cat, 'konto_rmk': konto_rmk}
            self._save_state()
        except Exception:
            pass

        for it in self.rmk_items:
            if cur_company and it.firma and it.firma != cur_company:
                continue
            if cat and cat != "Wszystkie":
                try:
                    if str(it.kategoria).strip().casefold() != cat.casefold():
                        continue
                except Exception:
                    continue
            if konto_rmk and konto_rmk != "Wszystkie":
                if str(it.konto_rmk).strip() != konto_rmk:
                    continue

            # Użyj ulepszonego algorytmu rozliczania miesięcznego
            monthly_amounts = calculate_monthly_amounts_improved(it.kwota, it.data_start, it.liczba_mies)
            
            for i in range(it.liczba_mies):
                mdate = month_add(it.data_start, i)
                y = mdate.year
                if y < y1 or y > y2:
                    continue
                k = it.kategoria or ""
                keys.add(k)
                agg.setdefault(k, {})
                
                # Użyj odpowiedniej kwoty z ulepszonego algorytmu
                amount_to_add = monthly_amounts[i] if i < len(monthly_amounts) else 0.0
                agg[k][str(y)] = agg[k].get(str(y), 0.0) + amount_to_add

        for w in self.rmk_by_year_frame.winfo_children():
            w.destroy()
        cols = ["Kategoria"] + years + ["Razem"]
        tree = ttk.Treeview(self.rmk_by_year_frame, columns=cols, show='headings')
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=120, anchor='e' if c != 'Kategoria' else 'w')
        tree.pack(fill=tk.BOTH, expand=True)
        
        # Dodaj obramowania do tabeli RMK wg lat
        configure_single_treeview_borders(tree)

        for k in sorted(keys):
            row = [k]
            total = 0.0
            for y in years:
                v = agg.get(k, {}).get(y, 0.0)
                total += v
                row.append(thousand_sep(v))
            row.append(thousand_sep(total))
            tree.insert('', 'end', values=row)

        # footer: suma per year
        footer = ["SUMA"]
        grand = 0.0
        for y in years:
            s = sum(agg.get(k, {}).get(y, 0.0) for k in keys)
            grand += s
            footer.append(thousand_sep(s))
        footer.append(thousand_sep(grand))
        tree.insert('', 'end', values=footer)

    def _build_tab_slownik(self, nb):
        tab = ttk.Frame(nb)
        tab.configure(style='Light.TFrame')  # Lekko szare tło
        nb.add(tab, text="Słowniki")
        topf = ttk.Frame(tab)
        topf.pack(fill=tk.X, padx=8, pady=6)
        lf = ttk.Labelframe(tab, text="Kategorie RMK", padding=6)
        lf.pack(fill=tk.X, padx=8, pady=6)
        btnf = ttk.Frame(lf)
        btnf.pack(anchor='e', pady=4)
        self._btn(btnf, text="+ Dodaj", command=self.add_category, bootstyle='success').pack(side=tk.LEFT, padx=4)
        ttk.Button(btnf, text="Edit Edytuj", command=self.edit_category).pack(side=tk.LEFT, padx=4)
        ttk.Button(btnf, text="Usuń", command=self.delete_category).pack(side=tk.LEFT, padx=4)
        
        # Frame ze scrollbarem dla kategorii
        cat_frame = ttk.Frame(lf)
        cat_frame.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)
        
        self.cat_tree = ttk.Treeview(cat_frame, columns=("kategoria",), show='headings', height=3, style='Dictionary.Treeview')
        cat_scrollbar = ttk.Scrollbar(cat_frame, orient="vertical", command=self.cat_tree.yview)
        self.cat_tree.configure(yscrollcommand=cat_scrollbar.set)
        
        self.cat_tree.heading("kategoria", text="Kategoria")
        self.cat_tree.pack(side="left", fill=tk.BOTH, expand=True)
        cat_scrollbar.pack(side="right", fill="y")
        
        # Dodaj obramowania do tabeli kategorii
        configure_single_treeview_borders(self.cat_tree)
        lf2 = ttk.Labelframe(tab, text="Konta księgowe", padding=6)
        lf2.pack(fill=tk.X, padx=8, pady=6)
        btnf2 = ttk.Frame(lf2)
        btnf2.pack(anchor='e', pady=4)
        self._btn(btnf2, text="+ Dodaj", command=self.add_account, bootstyle='success').pack(side=tk.LEFT, padx=4)
        ttk.Button(btnf2, text="Edit Edytuj", command=self.edit_account).pack(side=tk.LEFT, padx=4)
        ttk.Button(btnf2, text="Usuń", command=self.delete_account).pack(side=tk.LEFT, padx=4)
        
        # Frame ze scrollbarem dla kont księgowych
        acc_frame = ttk.Frame(lf2)
        acc_frame.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)
        
        self.acc_tree = ttk.Treeview(acc_frame, columns=("konto", "opis"), show='headings', height=3, style='Dictionary.Treeview')
        acc_scrollbar = ttk.Scrollbar(acc_frame, orient="vertical", command=self.acc_tree.yview)
        self.acc_tree.configure(yscrollcommand=acc_scrollbar.set)
        
        self.acc_tree.heading("konto", text="Konto")
        self.acc_tree.heading("opis", text="Opis")
        self.acc_tree.pack(side="left", fill=tk.BOTH, expand=True)
        acc_scrollbar.pack(side="right", fill="y")
        
        # Dodaj obramowania do tabeli kont
        configure_single_treeview_borders(self.acc_tree)
        lf3 = ttk.Labelframe(tab, text="Konta RMK", padding=6)
        lf3.pack(fill=tk.BOTH, expand=True, padx=8, pady=6)
        btnf3 = ttk.Frame(lf3)
        btnf3.pack(anchor='e', pady=4)
        self._btn(btnf3, text="+ Dodaj", command=self.add_rmk_account, bootstyle='success').pack(side=tk.LEFT, padx=4)
        ttk.Button(btnf3, text="Edit Edytuj", command=self.edit_rmk_account).pack(side=tk.LEFT, padx=4)
        ttk.Button(btnf3, text="Usuń", command=self.delete_rmk_account).pack(side=tk.LEFT, padx=4)
        
        # Frame ze scrollbarem dla kont RMK
        rmk_acc_frame = ttk.Frame(lf3)
        rmk_acc_frame.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)
        
        self.rmk_acc_tree = ttk.Treeview(rmk_acc_frame, columns=("konto", "opis"), show='headings', height=6, style='Dictionary.Treeview')
        rmk_acc_scrollbar = ttk.Scrollbar(rmk_acc_frame, orient="vertical", command=self.rmk_acc_tree.yview)
        self.rmk_acc_tree.configure(yscrollcommand=rmk_acc_scrollbar.set)
        
        self.rmk_acc_tree.heading("konto", text="Konto RMK")
        self.rmk_acc_tree.heading("opis", text="Opis")
        self.rmk_acc_tree.column("konto", width=120, minwidth=80)
        self.rmk_acc_tree.column("opis", width=250, minwidth=150)
        self.rmk_acc_tree.pack(side="left", fill=tk.BOTH, expand=True)
        rmk_acc_scrollbar.pack(side="right", fill="y")
        
        # Dodaj obramowania do tabeli kont RMK
        configure_single_treeview_borders(self.rmk_acc_tree)
        self.refresh_cat_tree()
        self.refresh_acc_tree()
        self.refresh_rmk_acc_tree()

    def _build_tab_admin(self, nb):
        if not self.current_user_admin:
            return
        tab = ttk.Frame(nb)
        tab.configure(style='Light.TFrame')  # Lekko szare tło
        nb.add(tab, text="Admin")
        ttk.Label(tab, text="Użytkownicy systemu", background=BRAND_COLOR_LIGHT).pack(anchor='w', padx=8, pady=6)
        uframe = ttk.Frame(tab)
        uframe.pack(fill=tk.X, padx=8, pady=4)
        btnf = ttk.Frame(uframe)
        btnf.pack(anchor='e', pady=4)
        self._btn(btnf, text="+ Dodaj", command=self.add_user, bootstyle='success').pack(side=tk.LEFT, padx=4)
        ttk.Button(btnf, text="Edit Edytuj", command=self.edit_user).pack(side=tk.LEFT, padx=4)
        ttk.Button(btnf, text="Usuń", command=self.delete_user).pack(side=tk.LEFT, padx=4)
        
        # Frame ze scrollbarem dla użytkowników
        user_frame = ttk.Frame(uframe)
        user_frame.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)
        
        self.user_tree = ttk.Treeview(user_frame, columns=("uzytkownik", "rola", "firmy"), show='headings', height=5, style='Admin.Treeview')
        user_scrollbar = ttk.Scrollbar(user_frame, orient="vertical", command=self.user_tree.yview)
        self.user_tree.configure(yscrollcommand=user_scrollbar.set)
        
        self.user_tree.heading("uzytkownik", text="Użytkownik")
        self.user_tree.heading("rola", text="Rola")
        self.user_tree.heading("firmy", text="Firmy")
        self.user_tree.pack(side="left", fill=tk.BOTH, expand=True)
        user_scrollbar.pack(side="right", fill="y")
        
        # Dodaj obramowania do tabeli użytkowników
        configure_single_treeview_borders(self.user_tree)
        self.refresh_user_tree()
        ttk.Label(tab, text="Firmy", background=BRAND_COLOR_LIGHT).pack(anchor='w', padx=8, pady=6)
        cframe = ttk.Frame(tab)
        cframe.pack(fill=tk.X, padx=8, pady=4)
        btnf2 = ttk.Frame(cframe)
        btnf2.pack(anchor='e', pady=4)
        self._btn(btnf2, text="+ Dodaj", command=self.add_company, bootstyle='success').pack(side=tk.LEFT, padx=4)
        ttk.Button(btnf2, text="Edit Edytuj", command=self.edit_company).pack(side=tk.LEFT, padx=4)
        ttk.Button(btnf2, text="Usuń", command=self.delete_company).pack(side=tk.LEFT, padx=4)
        
        # Frame ze scrollbarem dla firm
        company_frame = ttk.Frame(cframe)
        company_frame.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)
        
        self.company_tree = ttk.Treeview(company_frame, columns=("firma",), show='headings', height=5, style='Admin.Treeview')
        company_scrollbar = ttk.Scrollbar(company_frame, orient="vertical", command=self.company_tree.yview)
        self.company_tree.configure(yscrollcommand=company_scrollbar.set)
        
        self.company_tree.heading("firma", text="Firma")
        self.company_tree.pack(side="left", fill=tk.BOTH, expand=True)
        company_scrollbar.pack(side="right", fill="y")
        
        # Dodaj obramowania do tabeli firm
        configure_single_treeview_borders(self.company_tree)
        self.refresh_company_tree()

    def _build_tab_reports(self, nb):
        tab = ttk.Frame(nb)
        tab.configure(style='Light.TFrame')  # Lekko szare tło
        nb.add(tab, text="Raporty")
        toolbar = ttk.Frame(tab)
        toolbar.pack(fill=tk.X, padx=8, pady=6)
        ttk.Label(toolbar, text="Kategoria:").pack(side=tk.LEFT, padx=4)
        self.rep_cat_cb = ttk.Combobox(toolbar, values=["Wszystkie"] + self.categories, state='readonly')
        self.rep_cat_cb.current(0)
        self.rep_cat_cb.pack(side=tk.LEFT, padx=4)
        # restore last report filters
        last_rep = getattr(self, 'view_state', {}).get('report_filters', {})
        try:
            if last_rep.get('category'):
                vals = ["Wszystkie"] + self.categories
                if last_rep['category'] in vals:
                    self.rep_cat_cb.set(last_rep['category'])
        except Exception:
            pass
        ttk.Label(toolbar, text="Od (YYYY-MM):").pack(side=tk.LEFT, padx=4)
        self.rep_od = ttk.Entry(toolbar, width=8)
        self.rep_od.pack(side=tk.LEFT, padx=4)
        ttk.Label(toolbar, text="Do (YYYY-MM):").pack(side=tk.LEFT, padx=4)
        self.rep_do = ttk.Entry(toolbar, width=8)
        self.rep_do.pack(side=tk.LEFT, padx=4)
        self._btn(toolbar, text="Generuj raport", command=self.generate_report, bootstyle='primary').pack(side=tk.LEFT, padx=8)
        ttk.Button(toolbar, text="Eksportuj PDF", command=self.export_report_pdf).pack(side=tk.LEFT, padx=6)
        # Paging controls for months window (show max 12 months)
        self.report_window_size = 12
        self.report_months_window_start = 0
        self.rep_prev_btn = ttk.Button(toolbar, text="◀", command=lambda: self.shift_report_window(-1))
        self.rep_next_btn = ttk.Button(toolbar, text="▶", command=lambda: self.shift_report_window(1))
        self.rep_window_label = ttk.Label(toolbar, text="")
        # pack paging controls on the right
        self.rep_next_btn.pack(side=tk.RIGHT, padx=2)
        self.rep_prev_btn.pack(side=tk.RIGHT, padx=2)
        self.rep_window_label.pack(side=tk.RIGHT, padx=8)
        self.report_frame = ttk.Frame(tab)
        self.report_frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=6)
        self.report_tree = None

    def shift_report_window(self, delta: int):
        # Move the months window by delta * window_size? Here delta is number of pages (1 or -1)
        if not hasattr(self, 'report_months_full') or not self.report_months_full:
            return
        total = len(self.report_months_full)
        max_start = max(0, total - self.report_window_size)
        new_start = self.report_months_window_start + delta * self.report_window_size
        new_start = max(0, min(new_start, max_start))
        if new_start != self.report_months_window_start:
            self.report_months_window_start = new_start
            # regenerate report keeping same filters
            self.generate_report()

    # ---- refresh helpers ----
    def refresh_rmk_tree(self):
        if not hasattr(self, 'tree'):
            return
        self.tree.delete(*self.tree.get_children())
        # Show only items for the current company for non-admin users
        try:
            is_admin = bool(getattr(self, 'current_user_admin', False))
        except Exception:
            is_admin = False
        cur_company = getattr(self, 'current_company', '')
        for it in self.rmk_items:
            # always filter by current_company if set - show only items assigned to that company
            if cur_company and it.firma and it.firma != cur_company:
                continue
            # format kwota for display using thousand_sep
            kw = thousand_sep(it.kwota)
            status_sym = '✓' if getattr(it, 'harmonogram_generated', False) else '✗'
            uwagi_text = getattr(it, 'uwagi', '') or ''  # Obsługa starych danych bez pola uwagi
            self.tree.insert('', 'end', iid=str(it.id), values=(it.id, it.opis, it.data_start.isoformat(), it.data_koniec.isoformat() if it.data_koniec else "", it.liczba_mies, kw, it.kategoria, it.konto_kosztowe, it.konto_rmk, it.numer_faktury, it.kontrahent, status_sym, uwagi_text))
            # color row green if generated, red otherwise
            try:
                if getattr(it, 'harmonogram_generated', False):
                    self.tree.item(str(it.id), tags=('gen',))
                else:
                    self.tree.item(str(it.id), tags=('ungen',))
            except Exception:
                pass
        # tag styles (ttk Treeview doesn't support tag foreground/bg directly in all themes)
        try:
            self.tree.tag_configure('gen', background='#d4edda', foreground="#7BDD92")  # Wyrazisty zielony
            self.tree.tag_configure('ungen', background='#f8d7da', foreground="#e97f8a")  # Wyrazisty czerwony
        except Exception:
            pass
        
        # Wymóż odświeżenie kolorów po załadowaniu danych
        try:
            self.tree.after(50, lambda: self.tree.event_generate('<<TreeviewOpen>>'))
        except Exception:
            pass
        # update harmonogram item combobox values
        try:
            vals = [f"{it.id}: {it.kategoria} | {it.opis} | {it.data_start.strftime('%Y-%m-%d')} - {it.data_koniec.strftime('%Y-%m-%d') if it.data_koniec else 'N/A'}" for it in self.rmk_items if not cur_company or not it.firma or it.firma == cur_company]
            if hasattr(self, 'harmo_item_cb'):
                self.harmo_item_cb['values'] = vals
                if vals:
                    try:
                        self.harmo_item_cb.current(0)
                    except Exception:
                        pass
        except Exception:
            pass

    def refresh_cat_tree(self):
        self.cat_tree.delete(*self.cat_tree.get_children())
        for c in self.categories:
            self.cat_tree.insert('', 'end', values=(c,))

    def refresh_acc_tree(self):
        self.acc_tree.delete(*self.acc_tree.get_children())
        
        # Pokaż konta dla aktualnej firmy
        current_company = getattr(self, 'current_company', '')
        if current_company:
            accounts = self.get_accounts_for_company(current_company)
            for a in accounts:
                self.acc_tree.insert('', 'end', values=(a['konto'], a.get('opis', '')))
        else:
            # Fallback: pokaż wszystkie konta (stary format)
            for a in self.accounts:
                self.acc_tree.insert('', 'end', values=(a['konto'], a.get('opis', '')))

    def refresh_rmk_acc_tree(self):
        self.rmk_acc_tree.delete(*self.rmk_acc_tree.get_children())
        
        # Pokaż konta RMK dla aktualnej firmy
        current_company = getattr(self, 'current_company', '')
        if current_company:
            rmk_accounts = self.get_rmk_accounts_for_company(current_company)
            for r in rmk_accounts:
                self.rmk_acc_tree.insert('', 'end', values=(r['konto'], r.get('opis', '')))
        else:
            # Fallback: pokaż wszystkie konta RMK (stary format)
            for r in self.rmk_accounts:
                self.rmk_acc_tree.insert('', 'end', values=(r['konto'], r.get('opis', '')))

    def refresh_user_tree(self):
        self.user_tree.delete(*self.user_tree.get_children())
        for u, d in self.users.items():
            firms = ", ".join(d.get('companies', []))
            self.user_tree.insert('', 'end', iid=u, values=(u, 'Admin' if d['is_admin'] else 'Użytkownik', firms))

    def refresh_company_tree(self):
        self.company_tree.delete(*self.company_tree.get_children())
        for c in self.companies:
            self.company_tree.insert('', 'end', values=(c,))

    # ---- RMK CRUD ----
    def add_item(self):
        dlg = ItemDialog(self, None)
        if dlg.result:
            new_id = (max([it.id for it in self.rmk_items]) + 1) if self.rmk_items else 1
            r = dlg.result
            firma = r.get('firma') or (self.current_company if hasattr(self, 'current_company') else '')
            item = RMKItem(new_id, r['opis'], r['data_start'], r['liczba_mies'], r['kwota'], firma, r['kategoria'], r['konto_kosztowe'], r['konto_rmk'], r['numer_faktury'], r['kontrahent'], r.get('uwagi', ''), r['data_koniec'])
            # Uwagi już są w konstruktorze
            # item.uwagi = r.get('uwagi', '')
            # new item - no harmonogram generated yet
            item.harmonogram_generated = False
            item.harmonogram = []
            self.rmk_items.append(item)
            self.refresh_rmk_tree()
            self._save_state()
    def import_excel(self):
        if openpyxl is None:
            messagebox.showerror(APP_NAME, "Brak biblioteki openpyxl. Zainstaluj zależność (pip install openpyxl).")
            return
        path = filedialog.askopenfilename(title="Wybierz plik Excel", filetypes=[('Excel', '*.xlsx;*.xlsm;*.xltx;*.xltm'), ('All', '*.*')])
        if not path:
            return
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
        except Exception as e:
            messagebox.showerror(APP_NAME, f"Błąd otwarcia pliku: {e}")
            return
        # Expected (flexible) columns: firma/kategoria/data_start/data_koniec/kwota/konto_kosztowe/konto_rmk/kontrahent
        # Build a mapping from canonical column names to worksheet indices by normalizing headers
        try:
            header_cells = next(ws.iter_rows(min_row=1, max_row=1))
        except Exception:
            messagebox.showerror(APP_NAME, "Nie można odczytać nagłówka arkusza.")
            return
        mapping = {}
        for i, cell in enumerate(header_cells):
            raw = str(cell.value).strip() if cell.value is not None else ''
            norm = re.sub(r'[^a-z0-9]', '', raw.lower())
            # detect common header variants
            if 'firma' in norm:
                mapping['firma'] = i
            elif 'kategoria' in norm:
                mapping['kategoria'] = i
            elif 'start' in norm or 'datastart' in norm or norm in ('od','dataod'):
                mapping['od'] = i
            elif 'koniec' in norm or 'data_koniec' in norm or norm in ('do','datado'):
                mapping['do'] = i
            elif 'kwot' in norm or 'wartosc' in norm or 'warto' in norm:
                mapping['wartosc'] = i
            elif 'konto' in norm and ('koszt' in norm or 'kosztow' in norm or 'kosztow' in norm):
                mapping['konto_kosztowe'] = i
            elif 'rmk' in norm or ('konto' in norm and 'rmk' in norm):
                mapping['konto_rmk'] = i
            elif 'kontrah' in norm or 'kontrahent' in norm:
                mapping['kontrahent'] = i
            elif 'uwag' in norm or 'uwagi' in norm or 'opis' in norm:
                mapping['uwagi'] = i
            elif 'faktur' in norm or 'numer' in norm:
                mapping['faktura'] = i
            else:
                # fallbacks for other localized/underscored headers
                if norm in ('datastart','datastart'):
                    mapping.setdefault('od', i)
                elif norm in ('datakoniec','datakoniec'):
                    mapping.setdefault('do', i)

        # Require at least date range and amount to continue
        if not ('od' in mapping and 'do' in mapping and 'wartosc' in mapping):
            messagebox.showerror(APP_NAME, "Nie znaleziono wymaganych kolumn (data start, data koniec, kwota) w arkuszu.")
            return
        # collect preview rows first
        preview_rows = []
        def parse_date(v):
            if not v:
                return None
            if isinstance(v, datetime):
                return v.date()
            if isinstance(v, date):
                return v
            s = str(v).strip()
            # try common formats
            for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y", "%Y.%m.%d"):
                try:
                    return datetime.strptime(s, fmt).date()
                except Exception:
                    pass
            # sometimes Excel gives day.month.year without leading zeros or with other separators
            # try to replace commas and spaces
            s2 = s.replace(',', '.').replace(' ', '')
            for fmt in ("%d.%m.%Y", "%d-%m-%Y"):
                try:
                    return datetime.strptime(s2, fmt).date()
                except Exception:
                    pass
            # fallback: try yyyy-mm as month start
            try:
                return datetime.strptime(s + "-01", "%Y-%m-%d").date()
            except Exception:
                return None

        def parse_float(v):
            if v is None:
                return 0.0
            if isinstance(v, (int, float)):
                return float(v)
            s = str(v).strip().replace(' ', '').replace(',', '.')
            try:
                return float(s)
            except Exception:
                return 0.0

        for row in ws.iter_rows(min_row=2):
            try:
                get = lambda name: (row[mapping[name]].value if name in mapping and mapping[name] < len(row) else None)
                firma = (get('firma') or '')
                if isinstance(firma, str):
                    firma = firma.strip()
                else:
                    try:
                        firma = str(firma)
                    except Exception:
                        firma = ''
                kategoria = (get('kategoria') or '')
                kategoria = kategoria.strip() if isinstance(kategoria, str) else str(kategoria) if kategoria is not None else ''
                od_val = get('od')
                do_val = get('do')
                konto_kosztowe = (get('konto_kosztowe') or '')
                konto_kosztowe = konto_kosztowe.strip() if isinstance(konto_kosztowe, str) else str(konto_kosztowe) if konto_kosztowe is not None else ''
                konto_rmk = (get('konto_rmk') or '')
                konto_rmk = konto_rmk.strip() if isinstance(konto_rmk, str) else str(konto_rmk) if konto_rmk is not None else ''
                wartosc = parse_float(get('wartosc'))
                kontrahent = (get('kontrahent') or '')
                kontrahent = kontrahent.strip() if isinstance(kontrahent, str) else str(kontrahent) if kontrahent is not None else ''
                uwagi_excel = (get('uwagi') or '')
                uwagi_excel = uwagi_excel.strip() if isinstance(uwagi_excel, str) else str(uwagi_excel) if uwagi_excel is not None else ''
                faktura = (get('faktura') or '')
                faktura = faktura.strip() if isinstance(faktura, str) else str(faktura) if faktura is not None else ''
                od_d = parse_date(od_val)
                do_d = parse_date(do_val)
                if not od_d or not do_d:
                    continue
                months = (do_d.year - od_d.year) * 12 + (do_d.month - od_d.month) + 1
                kwota = float(wartosc)
                preview_rows.append({
                    'firma': firma,
                    'kategoria': kategoria,
                    'uwagi_excel': uwagi_excel,
                    'faktura': faktura,
                    'od': od_d,
                    'do': do_d,
                    'konto_kosztowe': konto_kosztowe,
                    'konto_rmk': konto_rmk,
                    'wartosc': kwota,
                    'kontrahent': kontrahent,
                    'months': months
                })
            except Exception:
                continue

        if not preview_rows:
            messagebox.showinfo(APP_NAME, "Nie znaleziono poprawnych wierszy do zaimportowania.")
            return

        # show preview and ask for confirmation
        ok = self.show_import_preview(preview_rows)
        if not ok:
            messagebox.showinfo(APP_NAME, "Import anulowany.")
            return

        # apply preview rows
        added = 0
        for pr in preview_rows:
            try:
                new_id = (max([it.id for it in self.rmk_items]) + 1) if self.rmk_items else 1
                firma = pr.get('firma') or (self.current_company if hasattr(self, 'current_company') else '')
                
                # Nowa logika importu:
                # - uwagi z Excela (kolumna "uwagi") → opis pozycji  
                # - kategoria z Excela → pole uwagi w aplikacji
                uwagi_excel = pr.get('uwagi_excel', '') or ''
                kategoria_excel = pr.get('kategoria', '') or ''
                
                # Opis: jeśli są uwagi w Excelu, użyj ich, inaczej domyślny format
                if uwagi_excel:
                    opis_pozycji = uwagi_excel
                else:
                    opis_pozycji = kategoria_excel if kategoria_excel else "pozycja z Excel"
                
                item = RMKItem(new_id, opis_pozycji, pr['od'], pr['months'], pr['wartosc'], firma, kategoria_excel, pr['konto_kosztowe'], pr['konto_rmk'], pr.get('faktura', ''), pr['kontrahent'], uwagi_excel, pr['do'])
                
                # Kategoria z Excela już jest w konstruktorze jako uwagi
                # item.uwagi = kategoria_excel
                
                # imported items start without generated harmonogram
                item.harmonogram_generated = False
                item.harmonogram = []
                self.rmk_items.append(item)
                # Nie dodawaj kategorii z Excela do słownika kategorii (teraz trafia do uwag)
                # if pr['kategoria'] and pr['kategoria'] not in self.categories:
                #     self.categories.append(pr['kategoria'])
                
                # Dodaj konta do odpowiedniej firmy
                if pr['konto_kosztowe'] and firma:
                    self.add_account_to_company(firma, pr['konto_kosztowe'], '')
                    # Zachowaj też w starym formacie dla kompatybilności
                    if not any(a['konto'] == pr['konto_kosztowe'] for a in self.accounts):
                        self.accounts.append({'konto': pr['konto_kosztowe'], 'opis': ''})
                        
                if pr['konto_rmk'] and firma:
                    self.add_rmk_account_to_company(firma, pr['konto_rmk'], '')
                    # Zachowaj też w starym formacie dla kompatybilności
                    if not any(r['konto'] == pr['konto_rmk'] for r in self.rmk_accounts):
                        self.rmk_accounts.append({'konto': pr['konto_rmk'], 'opis': ''})
                if pr['firma'] and pr['firma'] not in self.companies:
                    self.companies.append(pr['firma'])
                added += 1
            except Exception:
                continue

        self.refresh_rmk_tree()
        self.refresh_cat_tree()
        self.refresh_acc_tree()
        self.refresh_rmk_acc_tree()
        self.refresh_company_tree()
        self._save_state()
        messagebox.showinfo(APP_NAME, f"Zaimportowano {added} wierszy z pliku Excel.")

    def show_import_preview(self, rows: List[Dict]) -> bool:
        dlg = tk.Toplevel(self)
        dlg.title("Podgląd importu")
        dlg.transient(self)
        try:
            dlg.grab_set()
        except Exception:
            pass
        dlg.geometry("900x400")
        frame = ttk.Frame(dlg, padding=8)
        frame.pack(fill=tk.BOTH, expand=True)
        cols = ("firma", "kategoria", "uwagi_excel", "faktura", "od", "do", "konto_kosztowe", "konto_rmk", "wartosc", "kontrahent")
        tv = ttk.Treeview(frame, columns=cols, show='headings')
        for c in cols:
            # Ustaw polskie nagłówki dla podglądu importu
            if c == "firma":
                header_text = "Firma"
            elif c == "kategoria":
                header_text = "Kategoria"
            elif c == "uwagi_excel":
                header_text = "Uwagi"
            elif c == "faktura":
                header_text = "Faktura"
            elif c == "od":
                header_text = "Od"
            elif c == "do":
                header_text = "Do"
            elif c == "konto_kosztowe":
                header_text = "Konto kosztowe"
            elif c == "konto_rmk":
                header_text = "Konto RMK"
            elif c == "wartosc":
                header_text = "Wartość"
            elif c == "kontrahent":
                header_text = "Kontrahent"
            else:
                header_text = str(c)
            tv.heading(c, text=header_text)
            tv.column(c, width=110, anchor='w' if c in ('firma','kategoria','kontrahent','uwagi_excel','faktura') else 'e')
        for r in rows:
            tv.insert('', 'end', values=(r['firma'], r['kategoria'], r['od'].isoformat(), r['do'].isoformat(), r['konto_kosztowe'], r['konto_rmk'], thousand_sep(r['wartosc']), r['kontrahent']))
        tv.pack(fill=tk.BOTH, expand=True)
        
        # Dodaj obramowania do tabeli podglądu importu
        configure_single_treeview_borders(tv)
        ans = {'ok': False}
        btnf = ttk.Frame(frame)
        btnf.pack(fill=tk.X, pady=6)
        def on_confirm():
            ans['ok'] = True
            dlg.destroy()
        def on_cancel():
            ans['ok'] = False
            dlg.destroy()
        ttk.Button(btnf, text="Anuluj", command=on_cancel).pack(side=tk.RIGHT, padx=6)
        self._btn(btnf, text="Importuj", command=on_confirm, bootstyle='primary').pack(side=tk.RIGHT, padx=6)
        dlg.wait_window()
        return ans['ok']

    def edit_item(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo(APP_NAME, "Wybierz pozycję do edycji.")
            return
        iid = sel[0]
        it_id = int(iid)
        item = next((x for x in self.rmk_items if x.id == it_id), None)
        if not item:
            messagebox.showerror(APP_NAME, "Nie znaleziono pozycji.")
            return
        dlg = ItemDialog(self, item)
        if dlg.result:
            r = dlg.result
            # For non-admins firma should remain unchanged; admins may change it via dialog
            try:
                is_admin = bool(getattr(self, 'current_user_admin', False))
            except Exception:
                is_admin = False
            if is_admin:
                item.firma = r.get('firma') or item.firma
            item.opis = r['opis']
            item.data_start = r['data_start']
            item.liczba_mies = r['liczba_mies']
            item.kwota = r['kwota']
            item.kategoria = r['kategoria']
            item.konto_kosztowe = r['konto_kosztowe']
            item.konto_rmk = r['konto_rmk']
            item.numer_faktury = r['numer_faktury']
            item.kontrahent = r['kontrahent']
            item.uwagi = r.get('uwagi', '')
            item.data_koniec = r['data_koniec']
            # editing item invalidates previous harmonogram
            try:
                item.harmonogram_generated = False
                item.harmonogram = []
            except Exception:
                pass
            self.refresh_rmk_tree()
            self._save_state()

    def delete_item(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo(APP_NAME, "Wybierz pozycję do usunięcia.")
            return
        if messagebox.askyesno(APP_NAME, "Czy na pewno usunąć wybrane pozycje?"):
            for s in sel:
                try:
                    it_id = int(s)
                    self.rmk_items = [x for x in self.rmk_items if x.id != it_id]
                except:
                    pass
                self.tree.delete(s)
            self._save_state()

    # ---- harmonogram ----
    def generate_harmonogram(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo(APP_NAME, "Wybierz pozycję z listy RMK.")
            return
        iid = sel[0]
        vals = self.tree.item(iid, 'values')
        start = datetime.strptime(str(vals[2]), "%Y-%m-%d").date()
        end = datetime.strptime(str(vals[3]), "%Y-%m-%d").date()  # data końca
        liczba_mies = int(vals[4])
        # vals[5] may be formatted with thousand_sep (e.g. '1 234,56') -> normalize
        kw_str = str(vals[5])
        kw_str = kw_str.replace(' ', '').replace(',', '.')
        try:
            kwota = float(kw_str)
        except Exception:
            kwota = 0.0
            
        # Nowa logika: precyzyjne dzielenie według dni z transpozycją tabeli
        self.harmo_tree.delete(*self.harmo_tree.get_children())
        
        # Zbierz dane harmonogramu
        harmonogram_data = {}  # miesiac -> kwota
        harmonogram_rows = []
        
        # Oblicz całkowitą liczbę dni
        total_days = (end - start).days + 1
        kwota_na_dzien = kwota / total_days if total_days > 0 else 0
        
        current_date = start
        suma_kwot = 0.0
        months_list = []  # Lista miesięcy w kolejności
        
        while current_date <= end:
            # Znajdź ostatni dzień tego miesiąca w zakresie
            year = current_date.year
            month = current_date.month
            
            # Ostatni dzień miesiąca
            if month == 12:
                next_month_start = date(year + 1, 1, 1)
            else:
                next_month_start = date(year, month + 1, 1)
            last_day_of_month = next_month_start - timedelta(days=1)
            
            # Koniec okresu dla tego miesiąca (min z końca okresu i końca miesiąca)
            period_end = min(end, last_day_of_month)
            
            # Liczba dni w tym miesiącu dla naszego okresu
            days_in_period = (period_end - current_date).days + 1
            
            # Kwota dla tego miesiąca
            kwota_miesiac = round(kwota_na_dzien * days_in_period, 2)
            suma_kwot += kwota_miesiac
            
            month_key = current_date.strftime("%Y-%m")
            months_list.append(month_key)
            harmonogram_data[month_key] = kwota_miesiac
            
            harmonogram_rows.append({
                'miesiac': month_key, 
                'kwota': float(kwota_miesiac), 
                'konto': vals[7], 
                'konto_rmk': vals[8], 
                'kategoria': vals[6]
            })
            
            # Przejdź do pierwszego dnia następnego miesiąca
            current_date = next_month_start
        
        # Sprawdź różnicę i dodaj do ostatniego miesiąca jeśli potrzeba
        roznica = round(kwota - suma_kwot, 2)
        if abs(roznica) > 0.01 and months_list:  # Jeśli różnica > 1 grosz
            # Dodaj różnicę do ostatniego miesiąca
            last_month = months_list[-1]
            harmonogram_data[last_month] += roznica
            harmonogram_rows[-1]['kwota'] += roznica
        
        # Skonfiguruj kolumny tabeli (dynamicznie)
        # Kolumny: Label + wszystkie miesiące + Razem
        columns = ["Pozycja"] + months_list + ["Razem"]
        self.harmo_tree["columns"] = columns
        
        # Skonfiguruj nagłówki i szerokości kolumn
        for col in columns:
            if col == "Pozycja":
                self.harmo_tree.heading(col, text="Pozycja")
                self.harmo_tree.column(col, width=120, anchor='w')
            elif col == "Razem":
                self.harmo_tree.heading(col, text="Razem")
                self.harmo_tree.column(col, width=120, anchor='e')
            else:
                # Miesiące
                self.harmo_tree.heading(col, text=col)
                self.harmo_tree.column(col, width=100, anchor='e')
        
        # Dodaj jeden wiersz z kwotami
        row_values = ["Kwoty miesięczne"]
        total_sum = 0.0
        for month in months_list:
            kwota_m = harmonogram_data[month]
            row_values.append(thousand_sep(kwota_m))
            total_sum += kwota_m
        row_values.append(thousand_sep(total_sum))
        
        self.harmo_tree.insert('', 'end', values=row_values)

        # persist into RMKItem
        try:
            it_id = int(iid)
            item = next((x for x in self.rmk_items if x.id == it_id), None)
            if item:
                item.harmonogram = harmonogram_rows
                item.harmonogram_generated = True
                self._save_state()
                # refresh list to update status
                self.refresh_rmk_tree()
        except Exception:
            pass
        
        # Pokaż informacje o wygenerowanym harmonogramie
        miesiecy = len(harmonogram_rows)
        suma_kontrolna = sum(row['kwota'] for row in harmonogram_rows)
        messagebox.showinfo(APP_NAME, f"Wygenerowano harmonogram na {miesiecy} miesięcy.\nCałkowita kwota: {thousand_sep(kwota)}\nSuma harmonogramu: {thousand_sep(suma_kontrolna)}\nCałkowite dni: {total_days}")

    def show_selected_harmonogram(self):
        # show harmonogram for selected item from combobox
        sel = None
        try:
            sel = self.harmo_item_cb.get()
        except Exception:
            pass
        if not sel:
            messagebox.showinfo(APP_NAME, "Wybierz pozycję RMK z listy.")
            return
        # parse id from selection like '12: Opis...'
        try:
            it_id = int(str(sel).split(':', 1)[0])
        except Exception:
            messagebox.showerror(APP_NAME, "Niepoprawny wybór pozycji RMK.")
            return
        item = next((x for x in self.rmk_items if x.id == it_id), None)
        if not item:
            messagebox.showerror(APP_NAME, "Nie znaleziono pozycji RMK.")
            return
        
        # Sprawdź czy pozycja należy do aktualnej firmy
        cur_company = getattr(self, 'current_company', '')
        if cur_company and item.firma and item.firma != cur_company:
            messagebox.showwarning(APP_NAME, f"Ta pozycja RMK należy do firmy '{item.firma}', a obecnie wybrana jest firma '{cur_company}'.")
            return
            
        # if item has persisted harmonogram, show it; otherwise generate on the fly
        self.harmo_tree.delete(*self.harmo_tree.get_children())
        
        if getattr(item, 'harmonogram_generated', False) and getattr(item, 'harmonogram', None):
            # Pokaż zapisany harmonogram w nowym formacie (kolumny = miesiące)
            harmonogram_data = {}
            months_list = []
            
            for r in item.harmonogram:
                month = r.get('miesiac')
                kwota = r.get('kwota', 0.0)
                if month:
                    months_list.append(month)
                    harmonogram_data[month] = kwota
            
            # Skonfiguruj kolumny tabeli
            columns = ["Pozycja"] + months_list + ["Razem"]
            self.harmo_tree["columns"] = columns
            
            # Skonfiguruj nagłówki i szerokości kolumn
            for col in columns:
                if col == "Pozycja":
                    self.harmo_tree.heading(col, text="Pozycja")
                    self.harmo_tree.column(col, width=120, anchor='w')
                elif col == "Razem":
                    self.harmo_tree.heading(col, text="Razem")
                    self.harmo_tree.column(col, width=120, anchor='e')
                else:
                    self.harmo_tree.heading(col, text=col)
                    self.harmo_tree.column(col, width=100, anchor='e')
            
            # Dodaj wiersz z kwotami
            row_values = ["Kwoty miesięczne"]
            total_sum = 0.0
            for month in months_list:
                kwota_m = harmonogram_data.get(month, 0.0)
                row_values.append(thousand_sep(kwota_m))
                total_sum += kwota_m
            row_values.append(thousand_sep(total_sum))
            
            self.harmo_tree.insert('', 'end', values=row_values)
            messagebox.showinfo(APP_NAME, f"Pokażono zapisany harmonogram dla pozycji {item.id}.")
            return
            
        # otherwise build a temporary harmonogram using improved algorithm
        start = item.data_start
        liczba_mies = int(item.liczba_mies)
        kwota = float(item.kwota)
        
        # Użyj ulepszonego algorytmu
        monthly_amounts = calculate_monthly_amounts_improved(kwota, start, liczba_mies)
        
        months_list = []
        harmonogram_data = {}
        
        for i in range(liczba_mies):
            m = month_add(start, i)
            month_key = m.strftime("%Y-%m")
            part = monthly_amounts[i] if i < len(monthly_amounts) else 0.0
            months_list.append(month_key)
            harmonogram_data[month_key] = part
        
        # Skonfiguruj kolumny tabeli
        columns = ["Pozycja"] + months_list + ["Razem"]
        self.harmo_tree["columns"] = columns
        
        # Skonfiguruj nagłówki i szerokości kolumn
        for col in columns:
            if col == "Pozycja":
                self.harmo_tree.heading(col, text="Pozycja")
                self.harmo_tree.column(col, width=120, anchor='w')
            elif col == "Razem":
                self.harmo_tree.heading(col, text="Razem")
                self.harmo_tree.column(col, width=120, anchor='e')
            else:
                self.harmo_tree.heading(col, text=col)
                self.harmo_tree.column(col, width=100, anchor='e')
        
        # Dodaj wiersz z kwotami
        row_values = ["Kwoty miesięczne"]
        total_sum = 0.0
        for month in months_list:
            kwota_m = harmonogram_data[month]
            row_values.append(thousand_sep(kwota_m))
            total_sum += kwota_m
        row_values.append(thousand_sep(total_sum))
        
        self.harmo_tree.insert('', 'end', values=row_values)
        messagebox.showinfo(APP_NAME, f"Pokażono harmonogram dla pozycji {item.id} (wygenerowany tymczasowo).")

    def filter_harmonogram(self):
        dlg = tk.Toplevel(self)
        dlg.title("Filtr harmonogramu")
        dlg.transient(self)
        dlg.grab_set()
        ttk.Label(dlg, text="Data od (YYYY-MM):").grid(row=0, column=0, padx=6, pady=6, sticky='e')
        ttk.Label(dlg, text="Data do (YYYY-MM):").grid(row=1, column=0, padx=6, pady=6, sticky='e')
        ttk.Label(dlg, text="Kategoria (opcjonalnie):").grid(row=2, column=0, padx=6, pady=6, sticky='e')
        e1 = ttk.Entry(dlg)
        e2 = ttk.Entry(dlg)
        cb = ttk.Combobox(dlg, values=[""] + self.categories, state='readonly')
        e1.grid(row=0, column=1, padx=6, pady=6)
        e2.grid(row=1, column=1, padx=6, pady=6)
        cb.grid(row=2, column=1, padx=6, pady=6)
        btnf = ttk.Frame(dlg)
        btnf.grid(row=3, column=0, columnspan=2, pady=8)
        def apply_filter():
            od = e1.get().strip()
            do = e2.get().strip()
            kat = cb.get().strip()
            def parse_ym(s):
                if not s:
                    return None
                try:
                    return datetime.strptime(s + "-01", "%Y-%m-%d").date()
                except:
                    return None
            od_d = parse_ym(od)
            do_d = parse_ym(do)
            if (od and not od_d) or (do and not do_d):
                messagebox.showerror("Błąd", "Niepoprawny format daty. Użyj YYYY-MM.")
                return
            # reattach all first
            for iid in self.harmo_tree.get_children():
                try:
                    self.harmo_tree.reattach(iid, '', 'end')
                except Exception:
                    pass

            filter_by_date = bool(od_d or do_d)
            filter_by_cat = bool(kat)
            kat_norm = kat.strip().casefold() if filter_by_cat else ''

            for iid in self.harmo_tree.get_children():
                vals = self.harmo_tree.item(iid, 'values')
                # vals[0] -> month (YYYY-MM), vals[4] -> category
                # parse month to date for comparisons
                try:
                    mdt = datetime.strptime(str(vals[0]) + "-01", "%Y-%m-%d").date()
                except Exception:
                    # if cannot parse month, when filtering by date treat as non-matching
                    if filter_by_date:
                        self.harmo_tree.detach(iid)
                    continue

                # if filtering by date, drop rows outside range
                if filter_by_date:
                    if od_d and mdt < od_d:
                        self.harmo_tree.detach(iid)
                        continue
                    if do_d and mdt > do_d:
                        self.harmo_tree.detach(iid)
                        continue

                # if filtering by category, compare normalized strings
                if filter_by_cat:
                    try:
                        row_cat = str(vals[4]).strip().casefold()
                    except Exception:
                        row_cat = ''
                    if row_cat != kat_norm:
                        self.harmo_tree.detach(iid)
                        continue
            dlg.destroy()
        ttk.Button(btnf, text="Anuluj", command=dlg.destroy).pack(side=tk.RIGHT, padx=6)
        self._btn(btnf, text="Filtruj", command=apply_filter, bootstyle='primary').pack(side=tk.RIGHT, padx=6)
        dlg.wait_window()

    def aggregate_all_harmonogram(self):
        # Ask user which category to aggregate (default: Wszystkie)
        dlg = tk.Toplevel(self)
        dlg.title("Podsumowanie - wybierz kategorię")
        dlg.transient(self)
        try:
            dlg.grab_set()
        except Exception:
            pass
        ttk.Label(dlg, text="Kategoria:").grid(row=0, column=0, padx=6, pady=8, sticky='e')
        cat_cb = ttk.Combobox(dlg, values=["Wszystkie"] + self.categories, state='readonly')
        cat_cb.current(0)
        cat_cb.grid(row=0, column=1, padx=6, pady=8)
        btnf = ttk.Frame(dlg)
        btnf.grid(row=1, column=0, columnspan=2, pady=8)
        ans = {'ok': False, 'chosen': ''}
        def on_ok():
            try:
                ans['chosen'] = cat_cb.get()
            except Exception:
                ans['chosen'] = ''
            ans['ok'] = True
            dlg.destroy()
        def on_cancel():
            ans['ok'] = False
            ans['chosen'] = ''
            dlg.destroy()
        ttk.Button(btnf, text="Anuluj", command=on_cancel).pack(side=tk.RIGHT, padx=6)
        self._btn(btnf, text="Generuj", command=on_ok, bootstyle='primary').pack(side=tk.RIGHT, padx=6)
        dlg.wait_window()
        if not ans['ok']:
            return
        chosen = (ans.get('chosen') or '').strip()

        # Build aggregation respecting chosen category and current company for non-admins
        try:
            is_admin = bool(getattr(self, 'current_user_admin', False))
        except Exception:
            is_admin = False
        cur_company = getattr(self, 'current_company', '')

        agg: Dict[str, float] = {}
        for it in self.rmk_items:
            # always respect current company selection
            if cur_company and it.firma and it.firma != cur_company:
                continue
            # filter by chosen category unless 'Wszystkie'
            if chosen and chosen != "Wszystkie":
                try:
                    if str(it.kategoria).strip().casefold() != chosen.casefold():
                        continue
                except Exception:
                    continue
            # Użyj ulepszonego algorytmu rozliczania miesięcznego
            monthly_amounts = calculate_monthly_amounts_improved(it.kwota, it.data_start, it.liczba_mies)
            
            for i in range(it.liczba_mies):
                m = month_add(it.data_start, i).strftime("%Y-%m")
                part = monthly_amounts[i] if i < len(monthly_amounts) else 0.0
                agg[m] = agg.get(m, 0.0) + part

        self.harmo_tree.delete(*self.harmo_tree.get_children())
        for m in sorted(agg.keys()):
            self.harmo_tree.insert('', 'end', values=(m, thousand_sep(agg[m]), "", "", chosen if chosen else "Wszystkie"))
        messagebox.showinfo(APP_NAME, f"Wyświetlono podsumowanie RMK wg miesięcy (Kategoria: {chosen}).")

    # ---- słowniki CRUD ----
    def add_category(self):
        name = simpledialog.askstring(APP_NAME, "Nowa kategoria:")
        if name:
            if name in self.categories:
                messagebox.showinfo(APP_NAME, "Kategoria już istnieje.")
                return
            self.categories.append(name)
            self.refresh_cat_tree()
            self._save_state()

    def edit_category(self):
        sel = self.cat_tree.selection()
        if not sel:
            messagebox.showinfo(APP_NAME, "Wybierz kategorię do edycji.")
            return
        old = self.cat_tree.item(sel[0], 'values')[0]
        name = simpledialog.askstring(APP_NAME, "Nowa nazwa kategorii:", initialvalue=old)
        if name:
            idx = self.categories.index(old)
            self.categories[idx] = name
            self.refresh_cat_tree()
            self._save_state()

    def delete_category(self):
        sel = self.cat_tree.selection()
        if not sel:
            messagebox.showinfo(APP_NAME, "Wybierz kategorię do usunięcia.")
            return
        old = self.cat_tree.item(sel[0], 'values')[0]
        if messagebox.askyesno(APP_NAME, f"Usunąć kategorię '{old}'?"):
            self.categories = [c for c in self.categories if c != old]
            self.refresh_cat_tree()
            self._save_state()

    def add_account(self):
        dlg = AccountDialog(self, title="Nowe konto księgowe")
        if dlg.result:
            konto, opis = dlg.result
            current_company = getattr(self, 'current_company', '')
            if not current_company:
                messagebox.showinfo(APP_NAME, "Nie można dodać konta - brak wybranej firmy.")
                return
                
            # Sprawdź czy konto już istnieje dla tej firmy
            current_accounts = self.get_accounts_for_company(current_company)
            if any(a['konto'] == konto for a in current_accounts):
                messagebox.showinfo(APP_NAME, f"Konto {konto} już istnieje dla firmy {current_company}.")
                return
                
            # Dodaj konto do aktualnej firmy
            self.add_account_to_company(current_company, konto, opis)
            # Dodaj też do starego formatu dla kompatybilności
            if not any(a['konto'] == konto for a in self.accounts):
                self.accounts.append({"konto": konto, "opis": opis})
            
            self.refresh_acc_tree()
            self._save_state()

    def edit_account(self):
        sel = self.acc_tree.selection()
        if not sel:
            messagebox.showinfo(APP_NAME, "Wybierz konto do edycji.")
            return
        vals = self.acc_tree.item(sel[0], 'values')
        old_k, old_o = vals[0], vals[1]
        dlg = AccountDialog(self, title="Edytuj konto", konto=old_k, opis=old_o)
        if dlg.result:
            konto, opis = dlg.result
            for a in self.accounts:
                if a['konto'] == old_k:
                    a['konto'] = konto
                    a['opis'] = opis
                    break
            self.refresh_acc_tree()
            self._save_state()

    def delete_account(self):
        sel = self.acc_tree.selection()
        if not sel:
            messagebox.showinfo(APP_NAME, "Wybierz konto do usunięcia.")
            return
        old = self.acc_tree.item(sel[0], 'values')[0]
        if messagebox.askyesno(APP_NAME, f"Usunąć konto '{old}'?"):
            self.accounts = [a for a in self.accounts if a['konto'] != old]
            self.refresh_acc_tree()
            self._save_state()

    def add_rmk_account(self):
        dlg = RMKAccountDialog(self, title="Nowe konto RMK")
        if dlg.result:
            konto, opis = dlg.result
            current_company = getattr(self, 'current_company', '')
            if not current_company:
                messagebox.showinfo(APP_NAME, "Nie można dodać konta RMK - brak wybranej firmy.")
                return
                
            # Sprawdź czy konto już istnieje dla tej firmy
            current_rmk_accounts = self.get_rmk_accounts_for_company(current_company)
            if any(r['konto'] == konto for r in current_rmk_accounts):
                messagebox.showinfo(APP_NAME, f"Konto RMK {konto} już istnieje dla firmy {current_company}.")
                return
                
            # Dodaj konto do aktualnej firmy
            self.add_rmk_account_to_company(current_company, konto, opis)
            # Dodaj też do starego formatu dla kompatybilności
            if not any(r['konto'] == konto for r in self.rmk_accounts):
                self.rmk_accounts.append({"konto": konto, "opis": opis})
            
            self.refresh_rmk_acc_tree()
            self._save_state()

    def edit_rmk_account(self):
        sel = self.rmk_acc_tree.selection()
        if not sel:
            messagebox.showinfo(APP_NAME, "Wybierz konto RMK do edycji.")
            return
        vals = self.rmk_acc_tree.item(sel[0], 'values')
        old_k, old_o = vals[0], vals[1]
        dlg = RMKAccountDialog(self, title="Edytuj konto RMK", konto=old_k, opis=old_o)
        if dlg.result:
            konto, opis = dlg.result
            for r in self.rmk_accounts:
                if r['konto'] == old_k:
                    r['konto'] = konto
                    r['opis'] = opis
                    break
            self.refresh_rmk_acc_tree()
            self._save_state()

    def delete_rmk_account(self):
        sel = self.rmk_acc_tree.selection()
        if not sel:
            messagebox.showinfo(APP_NAME, "Wybierz konto RMK do usunięcia.")
            return
        old = self.rmk_acc_tree.item(sel[0], 'values')[0]
        if messagebox.askyesno(APP_NAME, f"Usunąć konto RMK '{old}'?"):
            self.rmk_accounts = [r for r in self.rmk_accounts if r['konto'] != old]
            self.refresh_rmk_acc_tree()
            self._save_state()

    # ---- admin CRUD ----
    def add_user(self):
        dlg = UserDialog(self)
        if dlg.result:
            r = dlg.result
            name = r['name']
            if name in self.users:
                messagebox.showinfo(APP_NAME, "Użytkownik już istnieje.")
                return
            self.users[name] = {'password': r['password'], 'is_admin': r['is_admin'], 'companies': r['companies']}
            self.refresh_user_tree()
            self._save_state()

    def edit_user(self):
        sel = self.user_tree.selection()
        if not sel:
            messagebox.showinfo(APP_NAME, "Wybierz użytkownika do edycji.")
            return
        username = sel[0]
        data = self.users.get(username, {})
        dlg = UserDialog(self, username, data)
        if dlg.result:
            r = dlg.result
            self.users[username]['password'] = r['password']
            self.users[username]['is_admin'] = r['is_admin']
            self.users[username]['companies'] = r['companies']
            self.refresh_user_tree()
            self._save_state()

    def delete_user(self):
        sel = self.user_tree.selection()
        if not sel:
            messagebox.showinfo(APP_NAME, "Wybierz użytkownika do usunięcia.")
            return
        username = sel[0]
        if username == self.current_user:
            messagebox.showerror(APP_NAME, "Nie można usunąć aktualnie zalogowanego użytkownika.")
            return
        if messagebox.askyesno(APP_NAME, f"Usunąć użytkownika '{username}'?"):
            if username in self.users:
                del self.users[username]
            self.refresh_user_tree()
            self._save_state()

    def add_company(self):
        dlg = CompanyDialog(self, title="Nowa firma")
        if dlg.result:
            name = dlg.result
            if name in self.companies:
                messagebox.showinfo(APP_NAME, "Firma już istnieje.")
                return
            self.companies.append(name)
            self.refresh_company_tree()
            self.refresh_user_tree()
            self.refresh_cat_tree()
            self._save_state()

    def edit_company(self):
        sel = self.company_tree.selection()
        if not sel:
            messagebox.showinfo(APP_NAME, "Wybierz firmę do edycji.")
            return
        old = self.company_tree.item(sel[0], 'values')[0]
        dlg = CompanyDialog(self, title="Edytuj firmę", name=old)
        if dlg.result:
            new = dlg.result
            idx = self.companies.index(old)
            self.companies[idx] = new
            self.refresh_company_tree()
            self.refresh_user_tree()
            self._save_state()

    def delete_company(self):
        sel = self.company_tree.selection()
        if not sel:
            messagebox.showinfo(APP_NAME, "Wybierz firmę do usunięcia.")
            return
        old = self.company_tree.item(sel[0], 'values')[0]
        if messagebox.askyesno(APP_NAME, f"Usunąć firmę '{old}'?"):
            self.companies = [c for c in self.companies if c != old]
            for u in self.users.values():
                comps = u.get('companies', [])
                if old in comps:
                    comps = [x for x in comps if x != old]
                    u['companies'] = comps
            self.refresh_company_tree()
            self.refresh_user_tree()
            self._save_state()

    def _change_company(self):
        # simple dialog to pick current company context
        dlg = tk.Toplevel(self)
        dlg.title("Zmień firmę")
        dlg.transient(self)
        try:
            dlg.grab_set()
        except Exception:
            pass
        ttk.Label(dlg, text="Firma:").grid(row=0, column=0, padx=6, pady=8, sticky='e')
        cb = ttk.Combobox(dlg, values=self.companies, state='readonly')
        cur = getattr(self, 'current_company', '')
        if cur and cur in self.companies:
            cb.set(cur)
        elif self.companies:
            cb.current(0)
        cb.grid(row=0, column=1, padx=6, pady=8)
        ans = {'ok': False}
        def on_ok():
            try:
                self.current_company = cb.get()
                # update title
                if self.current_company:
                    self.title(f"{APP_NAME} - {self.current_company}")
            except Exception:
                pass
            ans['ok'] = True
            dlg.destroy()
        def on_cancel():
            ans['ok'] = False
            dlg.destroy()
        btnf = ttk.Frame(dlg)
        btnf.grid(row=1, column=0, columnspan=2, pady=8)
        ttk.Button(btnf, text="Anuluj", command=on_cancel).pack(side=tk.RIGHT, padx=6)
        self._btn(btnf, text="Zmień", command=on_ok, bootstyle='primary').pack(side=tk.RIGHT, padx=6)
        dlg.wait_window()
        if ans['ok']:
            self.refresh_rmk_tree()

    # ---- raporty (kategorie w wierszach, miesiące w kolumnach) ----
    def generate_report(self):
        kat = self.rep_cat_cb.get()
        od = self.rep_od.get().strip()
        do = self.rep_do.get().strip()
        def parse_ym(s):
            if not s:
                return None
            try:
                return datetime.strptime(s + "-01", "%Y-%m-%d").date()
            except:
                return None
        od_d = parse_ym(od)
        do_d = parse_ym(do)
        if (od and not od_d) or (do and not do_d):
            messagebox.showerror("Błąd", "Niepoprawny format daty (użyj YYYY-MM) lub puste pola.")
            return
        data: Dict[str, Dict[str, float]] = {}
        cats = set(self.categories)
        try:
            is_admin = bool(getattr(self, 'current_user_admin', False))
        except Exception:
            is_admin = False
        cur_company = getattr(self, 'current_company', '')
        # persist report filters
        try:
            self.view_state = getattr(self, 'view_state', {})
            self.view_state['report_filters'] = {'category': kat, 'od': od, 'do': do}
            self._save_state()
        except Exception:
            pass

        for it in self.rmk_items:
            # always respect current company selection
            if cur_company and it.firma and it.firma != cur_company:
                continue
            # Użyj ulepszonego algorytmu rozliczania miesięcznego
            monthly_amounts = calculate_monthly_amounts_improved(it.kwota, it.data_start, it.liczba_mies)
            
            for i in range(it.liczba_mies):
                mdate = month_add(it.data_start, i)
                if od_d and mdate < od_d: continue
                if do_d and mdate > do_d: continue
                m = mdate.strftime("%Y-%m")
                part = monthly_amounts[i] if i < len(monthly_amounts) else 0.0
                data.setdefault(m, {})
                data[m][it.kategoria] = data[m].get(it.kategoria, 0.0) + part
                cats.add(it.kategoria)
        months = sorted(data.keys())
        cats = sorted(cats)
        if kat and kat != "Wszystkie":
            cats = [kat]
        for w in self.report_frame.winfo_children():
            w.destroy()

        # Prepare months window (max self.report_window_size months)
        self.report_months_full = months
        if not hasattr(self, 'report_months_window_start'):
            self.report_months_window_start = 0
        # clamp window start
        if self.report_months_window_start < 0:
            self.report_months_window_start = 0
        if self.report_months_window_start >= max(1, len(months)):
            self.report_months_window_start = max(0, len(months) - self.report_window_size)
        wnd_start = self.report_months_window_start
        wnd_end = wnd_start + self.report_window_size
        months_window = months[wnd_start:wnd_end]
        # update window label
        if months_window:
            lbl = f"Miesiące: {months_window[0]} - {months_window[-1]} ({min(len(months), self.report_window_size)}/{len(months)})"
        else:
            lbl = "Brak miesięcy"
        try:
            self.rep_window_label.config(text=lbl)
        except:
            pass

        # columns: kategoria | month1 | month2 | ... | Razem
        cols = ["kategoria"] + months_window + ["Razem"]
        tree_container, tree = self._make_scrolled_tree(self.report_frame, cols)
        # adjust headings text
        for c in cols:
            # Popraw nagłówek z obsługą polskich znaków
            if c == 'kategoria':
                header_text = 'Kategoria'
            elif c == 'Razem':
                header_text = 'Razem'
            else:
                # Dla miesięcy i innych kolumn
                header_text = str(c).replace('_', ' ')
                # Użyj title() zamiast capitalize() dla lepszej obsługi polskich znaków
                if header_text and len(header_text) > 1:
                    header_text = header_text[0].upper() + header_text[1:]
            tree.heading(c, text=header_text)
            tree.column(c, width=110, anchor='e' if c != 'kategoria' else 'w')
        tree_container.pack(fill=tk.BOTH, expand=True)

        # rows: one row per category
        sums_per_month = {m: 0.0 for m in months_window}
        grand_total = 0.0
        today = date.today()
        first_of_current_month = date(today.year, today.month, 1)
        for c in cats:
            row = [c]
            total = 0.0
            for m in months_window:
                v = data.get(m, {}).get(c, 0.0)
                total += v
                sums_per_month[m] += v
                row.append(thousand_sep(v))
            grand_total += total
            row.append(thousand_sep(total))
            tree.insert('', 'end', values=row)

        # footer: SUMA | sum(window months...) | grand
        if months_window:
            footer = ["SUMA"] + [thousand_sep(sums_per_month[m]) for m in months_window] + [thousand_sep(grand_total)]
            tree.insert('', 'end', values=footer)
        self.report_tree = tree
        # attach export path for convenience
        try:
            self.view_state = getattr(self, 'view_state', {})
            self.view_state['last_report_cols'] = cols
        except Exception:
            pass

    def export_rmk_next_year_pdf(self):
        # export current RMK-next-year view
        if not hasattr(self, 'rmk_year_frame'):
            messagebox.showinfo(APP_NAME, "Brak danych do eksportu.")
            return
        tree = self._find_tree_in(self.rmk_year_frame)
        if not tree:
            messagebox.showinfo(APP_NAME, "Brak tabeli do eksportu.")
            return
        f = filedialog.asksaveasfilename(defaultextension='.pdf', filetypes=[('PDF','*.pdf')])
        if not f: return
        ok = self.export_tree_to_pdf(tree, f"RMK - {self.rmk_year_cb.get()}", f)
        if ok:
            messagebox.showinfo(APP_NAME, "Zapisano PDF")

    def export_report_pdf(self):
        # export current report view
        if not hasattr(self, 'report_frame'):
            messagebox.showinfo(APP_NAME, "Brak danych do eksportu.")
            return
        tree = self._find_tree_in(self.report_frame)
        if not tree:
            messagebox.showinfo(APP_NAME, "Brak tabeli do eksportu.")
            return
        f = filedialog.asksaveasfilename(defaultextension='.pdf', filetypes=[('PDF','*.pdf')])
        if not f: return
        ok = self.export_tree_to_pdf(tree, "Raport RMK", f)
        if ok:
            messagebox.showinfo(APP_NAME, "Zapisano PDF")

    def _find_tree_in(self, parent):
        """Recursively search for first ttk.Treeview inside parent widget or its children.

        Returns the Treeview widget or None.
        """
        try:
            for widget in parent.winfo_children():
                if isinstance(widget, ttk.Treeview):
                    return widget
                # if it's a frame/container, search inside
                try:
                    found = self._find_tree_in(widget)
                    if found:
                        return found
                except Exception:
                    continue
        except Exception:
            return None
        return None

    def update_status(self, msg: str = ''):
        compile_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        self.status_var.set(f"{COMPANY_NAME} | {APP_NAME} {APP_VERSION} | Data kompilacji: {compile_date} {msg}")

    def _make_scrolled_tree(self, parent, columns):
        frame = ttk.Frame(parent)
        h = ttk.Scrollbar(frame, orient=tk.HORIZONTAL)
        v = ttk.Scrollbar(frame, orient=tk.VERTICAL)
        tree = ttk.Treeview(frame, columns=columns, show='headings', xscrollcommand=h.set, yscrollcommand=v.set)
        h.config(command=tree.xview)
        v.config(command=tree.yview)
        # configure headings - nie ustawiamy tutaj text, zostanie ustawiony w wywołującej funkcji
        for c in columns:
            tree.heading(c, text=str(c))  # tymczasowy tekst, zostanie nadpisany
            tree.column(c, width=120, anchor='e' if c != columns[0] else 'w')
        tree.grid(row=0, column=0, sticky='nsew')
        v.grid(row=0, column=1, sticky='ns')
        h.grid(row=1, column=0, sticky='ew')
        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)
        
        # Dodaj obramowania do tabeli raportów
        configure_single_treeview_borders(tree)
        
        return frame, tree

    def export_tree_to_pdf(self, tree: ttk.Treeview, title: str, filepath: str):
        if not REPORTLAB_AVAILABLE:
            messagebox.showerror(APP_NAME, "Brak biblioteki reportlab. Zainstaluj: pip install reportlab")
            return False
        # collect headers and rows
        cols = tree['columns']
        headers = [str(c) for c in cols]
        rows = []
        for iid in tree.get_children():
            vals = tree.item(iid, 'values')
            rows.append([str(v) for v in vals])

        try:
            # prepare document
            doc = SimpleDocTemplate(filepath, pagesize=landscape(A4), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=36)
            styles = getSampleStyleSheet()
            elems = []

            # ensure a TTF font is registered to support Polish characters
            font_name = 'DejaVuSans'
            font_registered = False
            try:
                from reportlab.pdfbase import pdfmetrics
                from reportlab.pdfbase.ttfonts import TTFont
                
                # Różne lokalizacje fontów dla różnych systemów
                candidates = []
                
                # Windows fonty
                if os.name == 'nt':
                    windows_fonts = [
                        os.path.join(os.environ.get('WINDIR', 'C:\\Windows'), 'Fonts', 'arial.ttf'),
                        os.path.join(os.environ.get('WINDIR', 'C:\\Windows'), 'Fonts', 'calibri.ttf'),
                        os.path.join(os.environ.get('WINDIR', 'C:\\Windows'), 'Fonts', 'DejaVuSans.ttf'),
                        'C:\\Windows\\Fonts\\arial.ttf',
                        'C:\\Windows\\Fonts\\calibri.ttf',
                    ]
                    candidates.extend(windows_fonts)
                
                # Linux fonty
                linux_fonts = [
                    '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
                    '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
                    '/usr/share/fonts/truetype/freefont/FreeSans.ttf',
                ]
                candidates.extend(linux_fonts)
                
                # Próbuj font z zasobów aplikacji (dla PyInstaller)
                try:
                    bundled_font = resource_path('fonts/DejaVuSans.ttf')
                    if os.path.exists(bundled_font):
                        candidates.insert(0, bundled_font)
                except Exception:
                    pass
                
                for p in candidates:
                    if os.path.exists(p):
                        try:
                            # Dla Windows użyj Arial jeśli jest dostępny
                            if 'arial' in p.lower():
                                font_name = 'Arial'
                            elif 'calibri' in p.lower():
                                font_name = 'Calibri'
                            else:
                                font_name = 'DejaVuSans'
                            
                            pdfmetrics.registerFont(TTFont(font_name, p))
                            font_registered = True
                            break
                        except Exception:
                            continue
                
                if not font_registered:
                    try:
                        pdfmetrics.registerFont(TTFont(font_name, 'DejaVuSans.ttf'))
                        font_registered = True
                    except Exception:
                        font_registered = False
                
                if not font_registered:
                    # fallback to built-in font (may not support Polish diacritics)
                    font_name = 'Helvetica'
                    
            except Exception:
                # if anything goes wrong, fall back to default
                font_name = 'Helvetica'

            # compute absolute base dir for reliable logo lookups (independent of CWD)
            base_dir = resource_path("")  # Użyj funkcji resource_path dla PyInstaller
            logo_path = None
            
            print(f"🔍 Debug logo - bazowy katalog: {base_dir}")
            
            try:
                logo_dir = os.path.join(base_dir, 'logo')
                print(f"🔍 Debug logo - katalog logo: {logo_dir}")
                print(f"🔍 Debug logo - katalog istnieje: {os.path.isdir(logo_dir)}")
                
                if os.path.isdir(logo_dir):
                    files = os.listdir(logo_dir)
                    print(f"🔍 Debug logo - pliki w katalogu: {files}")
                    
                    # Znajdź najlepszy plik logo (najwiekszy i prawidłowy)
                    best_logo = None
                    best_size = 0
                    
                    for fn in files:
                        print(f"🔍 Debug logo - sprawdzam plik: {fn}")
                        if fn.lower().endswith(('.png', '.jpg', '.jpeg')):
                            candidate_path = os.path.join(logo_dir, fn)
                            try:
                                # Sprawdź rozmiar pliku i czy da się załadować
                                file_size = os.path.getsize(candidate_path)
                                if file_size > 100:  # Min 100 bajtów
                                    # Test czy da się załadować jako obraz
                                    ir = ImageReader(candidate_path)
                                    iw, ih = ir.getSize()
                                    if iw > 10 and ih > 10:  # Min 10x10 pikseli
                                        print(f"OK Debug logo - KANDYDAT: {candidate_path} ({iw}x{ih}, {file_size}b)")
                                        if file_size > best_size:
                                            best_logo = candidate_path
                                            best_size = file_size
                                            print(f"🎯 Debug logo - NOWY NAJLEPSZY: {best_logo}")
                                    else:
                                        print(f"BŁĄD Debug logo - za małe wymiary: {candidate_path} ({iw}x{ih})")
                                else:
                                    print(f"BŁĄD Debug logo - za mały plik: {candidate_path} ({file_size}b)")
                            except Exception as e:
                                print(f"BŁĄD Debug logo - błąd ładowania: {candidate_path} - {e}")
                    
                    logo_path = best_logo
                    if logo_path:
                        print(f"OK Debug logo - WYBRANY FINALNIE: {logo_path}")
                        
                # fallback to common filenames in project root
                if not logo_path:
                    print("🔍 Debug logo - próbuję fallback w głównym katalogu")
                    for fn in ('logo.png', 'logo.jpg', 'logo.jpeg'):
                        p = os.path.join(base_dir, fn)
                        print(f"🔍 Debug logo - sprawdzam fallback: {p}")
                        if os.path.exists(p):
                            logo_path = p
                            print(f"OK Debug logo - ZNALEZIONO FALLBACK: {logo_path}")
                            break
            except Exception as e:
                print(f"BŁĄD Debug logo - błąd wyszukiwania: {e}")
                logo_path = None

            print(f"🎯 Debug logo - FINALNY WYNIK: {logo_path}")
            
            # Przygotuj logo dla nagłówka (mniejsze, po prawej stronie)
            logo_element = None
            if logo_path:
                try:
                    print(f"IMG Debug logo - próbuję załadować: {logo_path}")
                    ir = ImageReader(logo_path)
                    iw, ih = ir.getSize()
                    # Zmniejszone logo dla nagłówka - max 80px szerokości
                    max_w = 80
                    scale = min(1.0, float(max_w) / float(iw)) if iw and iw > 0 else 1.0
                    img_w = iw * scale
                    img_h = ih * scale
                    logo_element = Image(logo_path, width=img_w, height=img_h)
                    print(f"OK Debug logo - przygotowano dla nagłówka ({img_w}x{img_h})")
                except Exception as e:
                    print(f"BŁĄD Debug logo - błąd ładowania: {e}")
                    logo_element = None
            
            if not logo_element:
                print("BŁĄD Debug logo - brak logo w nagłówku PDF")

            # Header: app name + title + logo w tabeli (use registered font)
            title_style = styles.get('Title', styles['Normal']).clone('CustomTitle')
            normal_style = styles.get('Normal').clone('CustomNormal')
            try:
                title_style.fontName = font_name
                normal_style.fontName = font_name
            except Exception:
                pass
            
            # Utwórz nagłówek jako tabelę z tytułem po lewej i logo po prawej
            hdr = f"{APP_NAME} — {title}"
            header_paragraph = Paragraph(hdr, title_style)
            
            if logo_element:
                # Tabela z tytułem po lewej i logo po prawej
                header_data = [[header_paragraph, logo_element]]
                header_table = Table(header_data, colWidths=[400, 100])  # Szerokości kolumn
                header_table.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (0, 0), 'LEFT'),     # Tytuł po lewej
                    ('ALIGN', (1, 0), (1, 0), 'RIGHT'),    # Logo po prawej
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), # Wyśrodkowanie w pionie
                    ('LEFTPADDING', (0, 0), (-1, -1), 0),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                    ('TOPPADDING', (0, 0), (-1, -1), 0),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
                ]))
                elems.append(header_table)
            else:
                # Fallback: tylko tytuł bez logo
                elems.append(header_paragraph)
            elems.append(Spacer(1, 6))

            # Which data/filters were used
            comp = getattr(self, 'current_company', '') or ''
            year = ''
            try:
                year = self.rmk_year_cb.get().strip()
            except Exception:
                year = ''
            filters = []
            if comp:
                filters.append(f"Firma: {comp}")
            if year:
                filters.append(f"Rok: {year}")
            # If there is a status var, include its basic info (trim long text)
            status_text = ''
            try:
                status_text = getattr(self, 'status_var').get()
            except Exception:
                status_text = ''
            if status_text:
                filters.append(f"Status: {status_text}")

            if filters:
                elems.append(Paragraph("Dane użyte w raporcie: ", normal_style))
                elems.append(Paragraph(', '.join(filters), normal_style))
                elems.append(Spacer(1, 8))

            # Table with data
            # normalize headers (capitalize 'kategoria')
            headers = []
            for c in cols:
                hc = str(c)
                if hc.lower() == 'kategoria':
                    hc = 'Kategoria'
                headers.append(hc)
            data = [headers] + rows
            
            # Check if data contains Polish characters and warn if no proper font
            polish_chars = 'ąćęłńóśźżĄĆĘŁŃÓŚŹŻ'
            has_polish = False
            for row in data:
                for cell in row:
                    if any(char in str(cell) for char in polish_chars):
                        has_polish = True
                        break
                if has_polish:
                    break
            
            table = Table(data, repeatRows=1, hAlign='LEFT')
            # base table style
            tbl_style = [
                ('BACKGROUND', (0,0), (-1,0), colors.lightblue),
                ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('LEFTPADDING', (0,0), (-1,-1), 6),
                ('RIGHTPADDING', (0,0), (-1,-1), 6),
                # set font for entire table
                ('FONTNAME', (0,0), (-1,-1), font_name),
            ]
            # detect SUMA row (if last row starts with 'SUMA') and color it like the header
            try:
                if data and isinstance(data[-1], (list, tuple)) and str(data[-1][0]).strip().lower().startswith('suma'):
                    last_row_idx = len(data) - 1
                    tbl_style.append(('BACKGROUND', (0, last_row_idx), (-1, last_row_idx), colors.lightblue))
            except Exception:
                pass
            table.setStyle(TableStyle(tbl_style))
            elems.append(table)

            # Footer: generation info
            elems.append(Spacer(1, 12))
            gen_by = getattr(self, 'current_user', '') or ''
            gen_when = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            footer_lines = [f"Wygenerowano przez: {gen_by}" if gen_by else "",
                            f"Data generowania: {gen_when}",
                            f"{COMPANY_NAME} | {APP_NAME} {APP_VERSION}"]
            # add as small paragraphs with proper font
            small_style = normal_style.clone('CustomSmall')
            small_style.fontSize = 8
            for ln in footer_lines:
                if ln:
                    elems.append(Paragraph(ln, small_style))

            doc.build(elems)
            return True
        except Exception as e:
            messagebox.showerror(APP_NAME, f"Błąd eksportu PDF: {e}")
            return False

if __name__ == '__main__':
    try:
        print(">>> Uruchamiam aplikację RMK insGT...")
        
        # Sprawdź czy jesteśmy w exe i zapisz podstawowe info
        if getattr(sys, 'frozen', False):
            print("📦 Uruchamianie jako EXE")
            try:
                log_dir = os.path.join(os.path.expanduser("~"), "RMK_insGT")
                os.makedirs(log_dir, exist_ok=True)
                log_file = os.path.join(log_dir, 'debug.log')
                
                with open(log_file, 'w', encoding='utf-8') as f:
                    f.write(f"=== RMK insGT Debug Log ===\n")
                    f.write(f"Data: {datetime.now()}\n")
                    f.write(f"System: {os.name}\n")
                    f.write(f"Uruchamianie...\n\n")
                
                print(f"📝 Logi exe: {log_file}")
            except Exception as log_error:
                print(f"UWAGA Błąd logowania: {log_error}")
        else:
            print("Python Uruchamianie jako Python script")
        
        # Poprawka dla Windows - ustaw kodowanie stdout/stderr
        if os.name == 'nt':
            try:
                sys.stdout.reconfigure(encoding='utf-8')
                sys.stderr.reconfigure(encoding='utf-8')
            except Exception:
                pass
        
        app = RMKApp()
        app.mainloop()
    except Exception as e:
        # Zapisz błąd do pliku jeśli aplikacja się nie uruchomi
        try:
            import traceback
            error_log = f"RMK_error_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
            with open(error_log, 'w', encoding='utf-8') as f:
                f.write(f"Błąd uruchomienia aplikacji RMK insGT:\n")
                f.write(f"Data: {datetime.now()}\n")
                f.write(f"System: {os.name}\n")
                f.write(f"Python: {sys.version}\n")
                f.write(f"Błąd: {str(e)}\n\n")
                f.write("Traceback:\n")
                f.write(traceback.format_exc())
            
            # Pokaż komunikat użytkownikowi
            try:
                import tkinter.messagebox as mb
                mb.showerror("Błąd aplikacji", 
                           f"Wystąpił błąd podczas uruchamiania aplikacji.\n"
                           f"Szczegóły zapisano w pliku: {error_log}\n\n"
                           f"Błąd: {str(e)}")
            except Exception:
                print(f"BŁĄD: {e}")
                print(f"Log zapisano w: {error_log}")
        except Exception:
            print(f"Krytyczny błąd aplikacji: {e}")
        sys.exit(1)
